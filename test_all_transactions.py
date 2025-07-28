#!/usr/bin/env python3
"""
Comprehensive test script for all transaction types
Usage: python test_all_transactions.py
"""

import json
import requests
import time
from datetime import datetime
from copy import deepcopy

BASE_URL = "http://localhost:8000"

def print_header(text):
    """Print formatted header"""
    print(f"\n{'#'*60}")
    print(f"# {text}")
    print(f"{'#'*60}")

def print_result(response, show_full=True):
    """Print formatted response"""
    print(f"Status Code: {response.status_code}")
    
    if response.status_code == 200:
        data = response.json()
        print(f"Success: {data.get('success')}")
        
        if show_full:
            print("\nFull Response:")
            print(json.dumps(data, indent=2, default=str))
        else:
            # Summary only
            if data.get('service_response'):
                print(f"Policy GUID: {data['service_response'].get('policy_guid')}")
            if data.get('errors'):
                print(f"Errors: {data['errors']}")
    else:
        print(f"Error: {response.text}")

def load_base_payload():
    """Load base test payload"""
    with open("TEST.json", "r") as f:
        return json.load(f)

def test_bind_transaction():
    """Test Bind transaction"""
    print_header("Testing BIND Transaction")
    
    payload = load_base_payload()
    payload["transaction_type"] = "Bind"
    payload["transaction_id"] = f"test-bind-{int(time.time())}"
    
    print("Sending Bind transaction...")
    response = requests.post(f"{BASE_URL}/api/triton/transaction/new", json=payload)
    print_result(response)
    
    # Extract policy info for subsequent tests
    if response.status_code == 200:
        result = response.json()
        return result.get('service_response', {}).get('policy_guid')
    return None

def test_unbind_transaction(policy_number):
    """Test Unbind transaction"""
    print_header("Testing UNBIND Transaction")
    
    payload = load_base_payload()
    payload["transaction_type"] = "Unbind"
    payload["transaction_id"] = f"test-unbind-{int(time.time())}"
    payload["policy_number"] = policy_number
    
    print(f"Unbinding policy: {policy_number}")
    response = requests.post(f"{BASE_URL}/api/triton/transaction/new", json=payload)
    print_result(response)

def test_issue_transaction(policy_number):
    """Test Issue transaction"""
    print_header("Testing ISSUE Transaction")
    
    payload = load_base_payload()
    payload["transaction_type"] = "Issue"
    payload["transaction_id"] = f"test-issue-{int(time.time())}"
    payload["policy_number"] = policy_number
    
    print(f"Issuing policy: {policy_number}")
    response = requests.post(f"{BASE_URL}/api/triton/transaction/new", json=payload)
    print_result(response)

def test_endorsement_transaction(policy_number):
    """Test Midterm Endorsement transaction"""
    print_header("Testing MIDTERM ENDORSEMENT Transaction")
    
    payload = load_base_payload()
    payload["transaction_type"] = "Midterm Endorsement"
    payload["transaction_id"] = f"test-endorsement-{int(time.time())}"
    payload["policy_number"] = policy_number
    payload["midterm_endt_id"] = 12345
    payload["midterm_endt_description"] = "Test endorsement - coverage change"
    payload["midterm_endt_effective_from"] = "10/01/2025"
    payload["midterm_endt_endorsement_number"] = 1
    
    print(f"Creating endorsement for policy: {policy_number}")
    response = requests.post(f"{BASE_URL}/api/triton/transaction/new", json=payload)
    print_result(response)

def test_cancellation_transaction(policy_number):
    """Test Cancellation transaction"""
    print_header("Testing CANCELLATION Transaction")
    
    payload = load_base_payload()
    payload["transaction_type"] = "Cancellation"
    payload["transaction_id"] = f"test-cancel-{int(time.time())}"
    payload["policy_number"] = policy_number
    
    print(f"Cancelling policy: {policy_number}")
    response = requests.post(f"{BASE_URL}/api/triton/transaction/new", json=payload)
    print_result(response)

def test_excel_upload(transaction_id, quote_guid):
    """Test Excel rater upload"""
    print_header("Testing Excel Rater Upload")
    
    # Create a dummy Excel file
    import io
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Test Rater Data'
    ws['A2'] = 'Premium'
    ws['B2'] = 1000
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    files = {
        'file': ('test_rater.xlsx', excel_buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    }
    data = {
        'quote_guid': quote_guid
    }
    
    print(f"Uploading Excel rater for quote: {quote_guid}")
    response = requests.post(
        f"{BASE_URL}/api/triton/transaction/{transaction_id}/excel-rater",
        files=files,
        data=data
    )
    
    print(f"Status Code: {response.status_code}")
    if response.status_code == 200:
        print(f"Response: {response.json()}")
    else:
        print(f"Error: {response.text}")

def run_all_tests():
    """Run all transaction tests"""
    print("="*60)
    print("RSG Integration Service - Comprehensive Test Suite")
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*60)
    
    # Check service health
    print("\nChecking service health...")
    try:
        response = requests.get(f"{BASE_URL}/health")
        if response.status_code != 200:
            print("Service is not healthy!")
            return
        print("Service is healthy ✓")
    except:
        print("Service is not reachable! Make sure it's running.")
        return
    
    # Test sequence
    print("\n" + "="*60)
    print("STARTING TEST SEQUENCE")
    print("="*60)
    
    # 1. Test Bind (creates new policy)
    policy_guid = test_bind_transaction()
    
    if not policy_guid:
        print("\nBind failed - cannot continue with other tests")
        return
    
    # Use the original policy number for subsequent tests
    policy_number = load_base_payload()["policy_number"]
    
    # Wait a bit between tests
    time.sleep(2)
    
    # 2. Test Issue
    test_issue_transaction(policy_number)
    time.sleep(2)
    
    # 3. Test Midterm Endorsement
    test_endorsement_transaction(policy_number)
    time.sleep(2)
    
    # 4. Test Cancellation
    test_cancellation_transaction(policy_number)
    time.sleep(2)
    
    # 5. Test Unbind (should fail on cancelled policy)
    test_unbind_transaction(policy_number)
    
    print("\n" + "="*60)
    print("TEST SEQUENCE COMPLETE")
    print("="*60)

if __name__ == "__main__":
    # You can run individual tests or all tests
    import sys
    
    if len(sys.argv) > 1 and sys.argv[1] == "--all":
        run_all_tests()
    else:
        # Just run the basic bind test
        print("Running basic Bind test...")
        print("Use --all flag to run all transaction types")
        test_bind_transaction()