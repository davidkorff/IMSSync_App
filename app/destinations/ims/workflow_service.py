import logging
import os
import base64
from datetime import datetime, date
from typing import Dict, Any, Optional, Union, Tuple, List
import uuid
import json
import tempfile
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from app.models.transaction_models import (
    Transaction, TransactionStatus, IMSProcessingStatus,
    IMSInsured, IMSSubmission, IMSQuote, IMSPolicy
)
from app.services.ims_soap_client import IMSSoapClient
from app.core.config import settings

logger = logging.getLogger(__name__)

class IMSWorkflowService:
    """
    Handles the workflow for processing transactions through IMS
    """
    
    def __init__(self, environment=None):
        """Initialize with environment settings"""
        env = environment or settings.DEFAULT_ENVIRONMENT
        env_config = settings.IMS_ENVIRONMENTS.get(env)
        if not env_config:
            raise ValueError(f"Unknown environment: {env}")
        
        self.config_file = env_config["config_file"]
        self.username = env_config["username"]
        self.password = env_config["password"]
        self.soap_client = IMSSoapClient(self.config_file)
        self.token = None
        
        # Templates directory for Excel raters
        self.templates_dir = os.path.join(os.path.dirname(__file__), "../../templates")
        os.makedirs(self.templates_dir, exist_ok=True)
        
        # Cached lookup data
        self.business_types = None
        self.line_guids = None
        self.factor_sets = None
        self.rater_ids = None
    
    def authenticate(self) -> str:
        """Authenticate with IMS and get a token"""
        logger.info(f"Authenticating with IMS using {self.username}")
        self.token = self.soap_client.login(self.username, self.password)
        return self.token
    
    def process_transaction(self, transaction: Transaction) -> Transaction:
        """
        Process a transaction with IMS integration
        """
        logger.info(f"Processing transaction {transaction.transaction_id} in IMS")
        
        try:
            # Ensure we're authenticated
            if not self.token:
                self.authenticate()
            
            # Process based on transaction state
            if transaction.ims_processing.status == IMSProcessingStatus.PENDING:
                self._process_insured(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.INSURED_CREATED:
                self._process_submission(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.SUBMISSION_CREATED:
                self._process_quote(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.QUOTE_CREATED:
                self._process_rating(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.RATED:
                self._bind_policy(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.BOUND:
                self._issue_policy(transaction)
                
            # If we got here without errors, update the transaction status
            if transaction.ims_processing.status == IMSProcessingStatus.ISSUED:
                transaction.update_status(
                    TransactionStatus.COMPLETED,
                    f"Transaction successfully processed in IMS. Policy number: {transaction.ims_processing.policy.policy_number}"
                )
                
            return transaction
            
        except Exception as e:
            logger.error(f"Error processing transaction {transaction.transaction_id} in IMS: {str(e)}")
            transaction.update_status(
                TransactionStatus.FAILED,
                f"Error processing transaction in IMS: {str(e)}"
            )
            transaction.ims_processing.status = IMSProcessingStatus.ERROR
            return transaction
    
    def _process_insured(self, transaction: Transaction) -> None:
        """Process the insured data for the transaction"""
        transaction.ims_processing.add_log("Processing insured data")
        
        # Ensure we're authenticated
        if not self.soap_client.token:
            transaction.ims_processing.add_log(f"Authenticating with IMS using {self.username}")
            try:
                self.authenticate()
                transaction.ims_processing.add_log(f"Authentication successful")
            except Exception as e:
                transaction.ims_processing.add_log(f"Authentication failed: {str(e)}")
                raise
        
        # Extract insured data from transaction
        insured_data = self._extract_insured_data(transaction)
        
        # Check if insured exists
        try:
            transaction.ims_processing.add_log(f"Looking for existing insured: {insured_data['name']}")
            insured_guid = self.soap_client.find_insured_by_name(
                insured_data["name"], 
                insured_data.get("tax_id")
            )
            
            if insured_guid:
                transaction.ims_processing.add_log(f"Found existing insured with GUID: {insured_guid}")
            else:
                # Create new insured
                transaction.ims_processing.add_log(f"Creating new insured: {insured_data['name']}")
                insured_guid = self.soap_client.add_insured(insured_data)
                transaction.ims_processing.add_log(f"Created insured with GUID: {insured_guid}")
                
                # Add insured location if provided
                if transaction.parsed_data and "locations" in transaction.parsed_data and \
                   isinstance(transaction.parsed_data["locations"], list) and \
                   len(transaction.parsed_data["locations"]) > 0:
                    
                    location_data = transaction.parsed_data["locations"][0]
                    transaction.ims_processing.add_log(f"Adding location for insured: {location_data.get('address', '')}")
                    location_id = self.soap_client.add_insured_location(insured_guid, location_data)
                    transaction.ims_processing.add_log(f"Added location with ID: {location_id}")
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error processing insured: {str(e)}")
            raise
        
        # Store insured information
        transaction.ims_processing.insured = IMSInsured(
            guid=insured_guid,
            name=insured_data["name"],
            tax_id=insured_data.get("tax_id"),
            business_type_id=insured_data.get("business_type_id"),
            created_at=datetime.now()
        )
        
        # Update status
        transaction.ims_processing.update_status(
            IMSProcessingStatus.INSURED_CREATED,
            f"Insured processed successfully: {insured_guid}"
        )
    
    def _process_submission(self, transaction: Transaction) -> None:
        """Process the submission for the transaction"""
        transaction.ims_processing.add_log("Creating submission")
        
        # Extract submission data
        submission_data = self._extract_submission_data(transaction)
        
        # Add insured GUID from previous step
        submission_data["insured_guid"] = transaction.ims_processing.insured.guid
        
        try:
            # Create submission
            transaction.ims_processing.add_log(f"Creating submission for insured: {submission_data['insured_guid']}")
            submission_guid = self.soap_client.add_submission(submission_data)
            transaction.ims_processing.add_log(f"Created submission with GUID: {submission_guid}")
            
            # Store submission information
            transaction.ims_processing.submission = IMSSubmission(
                guid=submission_guid,
                insured_guid=submission_data["insured_guid"],
                submission_date=submission_data["submission_date"],
                producer_contact_guid=submission_data.get("producer_contact_guid"),
                underwriter_guid=submission_data.get("underwriter_guid"),
                producer_location_guid=submission_data.get("producer_location_guid"),
                created_at=datetime.now()
            )
            
            # Update status
            transaction.ims_processing.update_status(
                IMSProcessingStatus.SUBMISSION_CREATED,
                f"Submission created successfully: {submission_guid}"
            )
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error creating submission: {str(e)}")
            raise
    
    def _process_quote(self, transaction: Transaction) -> None:
        """Process the quote for the transaction"""
        transaction.ims_processing.add_log("Creating quote")
        
        # Extract quote data
        quote_data = self._extract_quote_data(transaction)
        
        # Add submission GUID from previous step
        quote_data["submission_guid"] = transaction.ims_processing.submission.guid
        
        try:
            # Create quote
            transaction.ims_processing.add_log(f"Creating quote for submission: {quote_data['submission_guid']}")
            quote_guid = self.soap_client.add_quote(quote_data)
            transaction.ims_processing.add_log(f"Created quote with GUID: {quote_guid}")
            
            # Store quote information
            transaction.ims_processing.quote = IMSQuote(
                guid=quote_guid,
                submission_guid=quote_data["submission_guid"],
                effective_date=quote_data["effective_date"],
                expiration_date=quote_data["expiration_date"],
                state=quote_data["state"],
                line_guid=quote_data.get("line_guid"),
                status_id=quote_data.get("status_id"),
                created_at=datetime.now()
            )
            
            # Update status
            transaction.ims_processing.update_status(
                IMSProcessingStatus.QUOTE_CREATED,
                f"Quote created successfully: {quote_guid}"
            )
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error creating quote: {str(e)}")
            raise
    
    def _process_rating(self, transaction: Transaction) -> None:
        """Process the rating for the transaction"""
        transaction.ims_processing.add_log("Processing rating data")
        
        # Determine which rating method to use
        if self._should_use_excel_rater(transaction):
            self._process_excel_rating(transaction)
        else:
            self._process_manual_rating(transaction)
        
        # Update status
        transaction.ims_processing.update_status(
            IMSProcessingStatus.RATED,
            "Quote rated successfully"
        )
    
    def _process_excel_rating(self, transaction: Transaction) -> None:
        """Process rating using Excel rater"""
        transaction.ims_processing.add_log("Using Excel rater")
        
        try:
            # Get the Excel template and populate it
            template_path = self._get_excel_template(transaction)
            if not template_path:
                raise ValueError("Excel template not found for this product/program")
            
            transaction.ims_processing.add_log(f"Using Excel template: {template_path}")
            
            # Create a populated Excel file
            populated_file_path = self._populate_excel_template(template_path, transaction)
            transaction.ims_processing.add_log(f"Populated Excel file created at: {populated_file_path}")
            
            # Store the path for future reference
            transaction.ims_processing.excel_rater_file_path = populated_file_path
            
            # Upload the Excel file to IMS
            transaction.ims_processing.add_log(f"Uploading Excel file to IMS for quote: {transaction.ims_processing.quote.guid}")
            rating_result = self._import_excel_rater(
                transaction.ims_processing.quote.guid,
                populated_file_path,
                transaction.ims_processing.excel_rater_id,
                transaction.ims_processing.factor_set_guid
            )
            
            # Store the rating results
            if transaction.ims_processing.quote:
                transaction.ims_processing.quote.premium = rating_result.get("premium_total")
                transaction.ims_processing.quote.quote_option_id = rating_result.get("quote_option_id")
            
            transaction.ims_processing.add_log(
                f"Excel rating completed. Premium: {rating_result.get('premium_total')}, "
                f"QuoteOptionID: {rating_result.get('quote_option_id')}"
            )
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error processing Excel rating: {str(e)}")
            raise
    
    def _process_manual_rating(self, transaction: Transaction) -> None:
        """Process rating manually by adding premiums"""
        transaction.ims_processing.add_log("Using manual rating")
        
        try:
            # Extract premium data from transaction
            premium_data = self._extract_premium_data(transaction)
            
            # Create quote option if needed
            if not transaction.ims_processing.quote.quote_option_id:
                transaction.ims_processing.add_log(f"Creating quote option for quote: {transaction.ims_processing.quote.guid}")
                option_id = self.soap_client.add_quote_option(transaction.ims_processing.quote.guid)
                transaction.ims_processing.quote.quote_option_id = option_id
                transaction.ims_processing.add_log(f"Created quote option: {option_id}")
            
            # Add premium to quote
            transaction.ims_processing.add_log(f"Adding premium: {premium_data.get('total_premium')}")
            self.soap_client.add_premium(
                transaction.ims_processing.quote.guid,
                transaction.ims_processing.quote.quote_option_id,
                premium_data.get("total_premium"),
                "Premium from external system"
            )
            
            # Update premium in quote
            transaction.ims_processing.quote.premium = premium_data.get("total_premium")
            transaction.ims_processing.add_log(
                f"Manual rating completed. Premium: {premium_data.get('total_premium')}"
            )
        
        except Exception as e:
            transaction.ims_processing.add_log(f"Error processing manual rating: {str(e)}")
            raise
    
    def _bind_policy(self, transaction: Transaction) -> None:
        """Bind the policy for the transaction"""
        transaction.ims_processing.add_log("Binding policy")
        
        if not transaction.ims_processing.quote.quote_option_id:
            raise ValueError("Cannot bind policy: Quote option ID not found")
        
        try:
            # Bind the policy
            transaction.ims_processing.add_log(f"Binding quote option: {transaction.ims_processing.quote.quote_option_id}")
            policy_number = self.soap_client.bind(transaction.ims_processing.quote.quote_option_id)
            transaction.ims_processing.add_log(f"Policy bound: {policy_number}")
            
            # Store policy information
            bound_date = datetime.now().date()
            if transaction.parsed_data and isinstance(transaction.parsed_data, dict):
                if transaction.parsed_data.get("bound_date"):
                    try:
                        bound_date = datetime.strptime(
                            transaction.parsed_data["bound_date"], "%Y-%m-%d"
                        ).date()
                    except ValueError:
                        pass
            
            transaction.ims_processing.policy = IMSPolicy(
                quote_guid=transaction.ims_processing.quote.guid,
                policy_number=policy_number,
                bound_date=bound_date,
                created_at=datetime.now()
            )
            
            # Update status
            transaction.ims_processing.update_status(
                IMSProcessingStatus.BOUND,
                f"Policy bound successfully: {policy_number}"
            )
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error binding policy: {str(e)}")
            raise
    
    def _issue_policy(self, transaction: Transaction) -> None:
        """Issue the policy for the transaction"""
        transaction.ims_processing.add_log("Issuing policy")
        
        try:
            # Issue the policy
            if transaction.ims_processing.policy and transaction.ims_processing.policy.policy_number:
                policy_number = transaction.ims_processing.policy.policy_number
                transaction.ims_processing.add_log(f"Issuing policy: {policy_number}")
                result = self.soap_client.issue_policy(policy_number)
                
                if result:
                    transaction.ims_processing.add_log(f"Policy issued successfully: {policy_number}")
                    
                    # Update status
                    transaction.ims_processing.update_status(
                        IMSProcessingStatus.ISSUED,
                        f"Policy issued successfully: {policy_number}"
                    )
                else:
                    transaction.ims_processing.add_log(f"Policy issue failed: {policy_number}")
                    raise ValueError(f"Failed to issue policy: {policy_number}")
                    
        except Exception as e:
            transaction.ims_processing.add_log(f"Error issuing policy: {str(e)}")
            raise
    
    def _import_excel_rater(self, quote_guid: str, excel_file_path: str, 
                           rater_id: int, factor_set_guid: str) -> Dict[str, Any]:
        """Import Excel rater to IMS"""
        try:
            logger.info(f"Importing Excel rater for quote: {quote_guid}")
            
            # Read the file and convert to base64
            with open(excel_file_path, "rb") as file:
                file_bytes = base64.b64encode(file.read()).decode("utf-8")
            
            # Call the IMS API to import the Excel rater
            result = self.soap_client.import_excel_rater(
                quote_guid, 
                file_bytes, 
                os.path.basename(excel_file_path), 
                rater_id, 
                factor_set_guid, 
                True
            )
            
            if result['success']:
                logger.info(f"Successfully imported Excel rater")
                
                # Find the first premium option
                if result['premiums'] and len(result['premiums']) > 0:
                    premium = result['premiums'][0]
                    return {
                        "success": True,
                        "premium_total": premium['premium_total'],
                        "fee_total": premium['fee_total'],
                        "quote_option_id": premium['quote_option_id']
                    }
                
                return {
                    "success": True,
                    "premium_total": 0.0,
                    "fee_total": 0.0,
                    "quote_option_id": None
                }
            else:
                logger.error(f"Failed to import Excel rater: {result.get('error_message')}")
                return {
                    "success": False,
                    "error_message": result.get('error_message')
                }
                
        except Exception as e:
            logger.error(f"Error importing Excel rater: {str(e)}")
            raise
    
    # Helper methods for data extraction and transformation
    
    def _extract_insured_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract insured data from transaction"""
        # This method would extract and transform insured data from the transaction
        # to match IMS API requirements
        
        if not transaction.parsed_data:
            raise ValueError("Cannot extract insured data: Missing parsed data")
        
        data = transaction.parsed_data
        result = {
            "name": data.get("insured", {}).get("name", "Unknown Insured"),
            "tax_id": data.get("insured", {}).get("tax_id"),
            "business_type_id": 1  # Default business type ID
        }
        
        # Extract contact information if available
        if "insured" in data and "contact" in data["insured"]:
            contact = data["insured"]["contact"]
            result.update({
                "contact_name": contact.get("name"),
                "contact_email": contact.get("email"),
                "contact_phone": contact.get("phone"),
                "address": contact.get("address"),
                "city": contact.get("city"),
                "state": contact.get("state"),
                "zip_code": contact.get("zip_code")
            })
        
        return result
    
    def _extract_submission_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract submission data from transaction"""
        # This method would extract and transform submission data from the transaction
        # to match IMS API requirements
        
        if not transaction.parsed_data:
            raise ValueError("Cannot extract submission data: Missing parsed data")
        
        data = transaction.parsed_data
        
        # Default to today's date if not provided
        submission_date = datetime.now().date()
        if data.get("bound_date"):
            try:
                submission_date = datetime.strptime(data["bound_date"], "%Y-%m-%d").date()
            except ValueError:
                pass
        
        result = {
            "insured_guid": None,  # Will be filled in by the caller
            "submission_date": submission_date,
            "producer_contact_guid": "00000000-0000-0000-0000-000000000000",  # Default
            "underwriter_guid": "00000000-0000-0000-0000-000000000000",  # Default
            "producer_location_guid": "00000000-0000-0000-0000-000000000000",  # Default
        }
        
        # Extract producer information if available
        if "producer" in data:
            producer = data["producer"]
            # In a real implementation, you would look up these GUIDs based on producer info
            # result["producer_contact_guid"] = self._get_producer_contact_guid(producer)
        
        # Extract underwriter information if available
        if data.get("underwriter"):
            # In a real implementation, you would look up the GUID based on underwriter name
            # result["underwriter_guid"] = self._get_underwriter_guid(data["underwriter"])
            pass
        
        return result
    
    def _extract_quote_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract quote data from transaction"""
        # This method would extract and transform quote data from the transaction
        # to match IMS API requirements
        
        if not transaction.parsed_data:
            raise ValueError("Cannot extract quote data: Missing parsed data")
        
        data = transaction.parsed_data
        
        # Parse dates
        effective_date = datetime.now().date()
        expiration_date = datetime.now().date().replace(year=datetime.now().year + 1)
        
        if data.get("effective_date"):
            try:
                effective_date = datetime.strptime(data["effective_date"], "%Y-%m-%d").date()
            except ValueError:
                pass
                
        if data.get("expiration_date"):
            try:
                expiration_date = datetime.strptime(data["expiration_date"], "%Y-%m-%d").date()
            except ValueError:
                pass
        
        result = {
            "submission_guid": None,  # Will be filled in by the caller
            "effective_date": effective_date,
            "expiration_date": expiration_date,
            "state": data.get("state", "TX"),  # Default to TX if not provided
            "line_guid": "00000000-0000-0000-0000-000000000000",  # Default
            "status_id": 1,  # Default status (New)
            "billing_type_id": 1,  # Default billing type (Agency Bill)
            "line_of_business": data.get("line_of_business", "General Liability")
        }
        
        # Extract line of business information if available
        if data.get("line_of_business"):
            # In a real implementation, you would look up the GUID based on line of business name
            # result["line_guid"] = self._get_line_guid(data["line_of_business"])
            pass
        
        return result
    
    def _extract_premium_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract premium data from transaction"""
        # This method would extract and transform premium data from the transaction
        # to match IMS API requirements
        
        if not transaction.parsed_data:
            raise ValueError("Cannot extract premium data: Missing parsed data")
        
        data = transaction.parsed_data
        
        # Default premium
        total_premium = 0.0
        
        # Extract premium amount
        if data.get("premium"):
            try:
                total_premium = float(data["premium"])
            except (ValueError, TypeError):
                pass
        
        # Extract coverages if available
        coverages = []
        if "coverages" in data and isinstance(data["coverages"], list):
            for coverage in data["coverages"]:
                if coverage.get("premium"):
                    try:
                        cov_premium = float(coverage["premium"])
                        coverages.append({
                            "type": coverage.get("type", "Unknown"),
                            "premium": cov_premium,
                            "limit": coverage.get("limit"),
                            "deductible": coverage.get("deductible")
                        })
                    except (ValueError, TypeError):
                        pass
        
        return {
            "total_premium": total_premium,
            "coverages": coverages
        }
    
    def _should_use_excel_rater(self, transaction: Transaction) -> bool:
        """Determine if the transaction should use Excel rater"""
        # This method would determine if the transaction should use Excel rater
        # based on the line of business, program, etc.
        
        # For now, always return True to use Excel rater
        return self._get_excel_template(transaction) is not None
    
    def _get_excel_template(self, transaction: Transaction) -> Optional[str]:
        """Get the Excel template for the transaction"""
        # This method would determine the appropriate Excel template
        # based on the line of business, program, etc.
        
        if not transaction.parsed_data:
            return None
        
        data = transaction.parsed_data
        line_of_business = data.get("line_of_business", "").lower().replace(" ", "_")
        program = data.get("program", "").lower().replace(" ", "_")
        state = data.get("state", "").upper()
        
        # Generate possible template paths
        template_paths = [
            f"{line_of_business}_{program}_{state}.xlsx",
            f"{line_of_business}_{program}.xlsx",
            f"{line_of_business}_{state}.xlsx",
            f"{line_of_business}.xlsx",
            "default.xlsx"
        ]
        
        # Check if any of the templates exist
        for template_name in template_paths:
            template_path = os.path.join(self.templates_dir, template_name)
            if os.path.exists(template_path):
                # Also set rater ID and factor set GUID based on template
                self._set_rater_info(transaction, template_name)
                return template_path
        
        # Create a sample template if none exists
        default_template = os.path.join(self.templates_dir, "default.xlsx")
        if not os.path.exists(default_template):
            self._create_sample_template(default_template)
            # Set default rater info
            self._set_rater_info(transaction, "default.xlsx")
            return default_template
            
        return None
    
    def _set_rater_info(self, transaction: Transaction, template_name: str) -> None:
        """Set rater ID and factor set GUID based on template"""
        # In a real implementation, you would look up these values based on the template
        transaction.ims_processing.excel_rater_id = 1
        transaction.ims_processing.factor_set_guid = "00000000-0000-0000-0000-000000000000"
    
    def _create_sample_template(self, template_path: str) -> None:
        """Create a sample Excel template"""
        # Create a basic Excel workbook with some sample data
        wb = Workbook()
        ws = wb.active
        ws.title = "Rate Calculator"
        
        # Add some headers
        headers = ["Policy_Number", "Effective_Date", "Expiration_Date", "Line_of_Business", 
                   "State", "Limit", "Deductible", "Premium"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add some formulas for premium calculation
        ws.cell(row=2, column=8, value="=IF(F2>0,F2*0.05,0)")  # Simple premium calculation
        
        # Save the workbook
        wb.save(template_path)
        logger.info(f"Created sample Excel template at {template_path}")
    
    def _populate_excel_template(self, template_path: str, transaction: Transaction) -> str:
        """Populate Excel template with transaction data"""
        if not transaction.parsed_data:
            raise ValueError("Cannot populate Excel template: Missing parsed data")
        
        data = transaction.parsed_data
        
        # Create a copy of the template in a temp file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_path = temp_file.name
        
        # Load the template
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Find headers
        headers = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers[header] = col
        
        # Populate basic data (row 2)
        if "Policy_Number" in headers and data.get("policy_number"):
            ws.cell(row=2, column=headers["Policy_Number"], value=data["policy_number"])
            
        if "Effective_Date" in headers and data.get("effective_date"):
            try:
                eff_date = datetime.strptime(data["effective_date"], "%Y-%m-%d").date()
                ws.cell(row=2, column=headers["Effective_Date"], value=eff_date)
            except ValueError:
                pass
                
        if "Expiration_Date" in headers and data.get("expiration_date"):
            try:
                exp_date = datetime.strptime(data["expiration_date"], "%Y-%m-%d").date()
                ws.cell(row=2, column=headers["Expiration_Date"], value=exp_date)
            except ValueError:
                pass
                
        if "Line_of_Business" in headers and data.get("line_of_business"):
            ws.cell(row=2, column=headers["Line_of_Business"], value=data["line_of_business"])
            
        if "State" in headers and data.get("state"):
            ws.cell(row=2, column=headers["State"], value=data["state"])
        
        # Populate coverage data
        if "coverages" in data and isinstance(data["coverages"], list) and data["coverages"]:
            coverage = data["coverages"][0]  # Just use the first coverage for now
            
            if "Limit" in headers and coverage.get("limit"):
                ws.cell(row=2, column=headers["Limit"], value=float(coverage["limit"]))
                
            if "Deductible" in headers and coverage.get("deductible"):
                ws.cell(row=2, column=headers["Deductible"], value=float(coverage["deductible"]))
        
        # Save the populated workbook
        wb.save(temp_path)
        
        return temp_path