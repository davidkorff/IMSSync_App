"""
IMS Workflow Orchestrator

This orchestrator coordinates the various IMS services to process complete workflows
such as creating a policy from scratch or updating existing policies.
"""

import logging
from typing import Dict, Any, Optional
from datetime import datetime, date
import os
import tempfile

from app.models.transaction_models import (
    Transaction, TransactionStatus, IMSProcessingStatus,
    IMSInsured, IMSSubmission, IMSQuote, IMSPolicy
)
from app.services.ims import (
    IMSInsuredService, IMSProducerService, IMSQuoteService,
    IMSDocumentService, IMSDataAccessService
)
from app.services.ims.field_mappings import get_mapper, FieldDestination
from app.core.config import settings

logger = logging.getLogger(__name__)


class IMSWorkflowOrchestrator:
    """
    Orchestrates IMS workflows using modular services
    """
    
    def __init__(self, environment: Optional[str] = None):
        """Initialize the orchestrator with all necessary services"""
        self.environment = environment or settings.DEFAULT_ENVIRONMENT
        
        # Initialize all services
        self.insured_service = IMSInsuredService(self.environment)
        self.producer_service = IMSProducerService(self.environment)
        self.quote_service = IMSQuoteService(self.environment)
        self.document_service = IMSDocumentService(self.environment)
        self.data_access_service = IMSDataAccessService(self.environment)
        
        # Templates directory for Excel raters
        self.templates_dir = os.path.join(os.path.dirname(__file__), "../../../templates")
        os.makedirs(self.templates_dir, exist_ok=True)
        
        logger.info(f"IMSWorkflowOrchestrator initialized for environment: {self.environment}")
    
    def process_transaction(self, transaction: Transaction) -> Transaction:
        """
        Process a transaction through the complete IMS workflow
        
        Args:
            transaction: The transaction to process
            
        Returns:
            The updated transaction
        """
        logger.info(f"Processing transaction {transaction.transaction_id} through IMS workflow")
        
        try:
            # Process based on transaction state
            if transaction.ims_processing.status == IMSProcessingStatus.PENDING:
                self._process_insured(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.INSURED_CREATED:
                self._process_submission(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.SUBMISSION_CREATED:
                self._process_quote(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.QUOTE_CREATED:
                self._process_rating(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.RATED:
                self._bind_policy(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.BOUND:
                self._issue_policy(transaction)
                
            # Update final status
            if transaction.ims_processing.status == IMSProcessingStatus.ISSUED:
                transaction.update_status(
                    TransactionStatus.COMPLETED,
                    f"Successfully processed in IMS. Policy: {transaction.ims_processing.policy.policy_number}"
                )
                
            return transaction
            
        except Exception as e:
            self._handle_workflow_error(transaction, e)
            return transaction
    
    def _process_insured(self, transaction: Transaction) -> None:
        """Process the insured step of the workflow"""
        transaction.ims_processing.add_log("Processing insured data")
        
        # Extract insured data
        insured_data = self._extract_insured_data(transaction)
        
        # Find or create insured
        insured_guid = self.insured_service.find_or_create_insured(insured_data)
        transaction.ims_processing.add_log(f"Insured GUID: {insured_guid}")
        
        # Add location if available
        if self._has_location_data(transaction):
            location_data = self._extract_location_data(transaction)
            location_id = self.insured_service.add_insured_location(insured_guid, location_data)
            transaction.ims_processing.add_log(f"Added location ID: {location_id}")
        
        # Handle additional insureds if present
        if self._has_additional_insureds(transaction):
            self._process_additional_insureds(transaction, insured_guid)
        
        # Store insured information
        transaction.ims_processing.insured = IMSInsured(
            guid=insured_guid,
            name=insured_data["name"],
            tax_id=insured_data.get("tax_id"),
            business_type_id=insured_data.get("business_type_id"),
            created_at=datetime.now()
        )
        
        # Update status
        transaction.ims_processing.update_status(
            IMSProcessingStatus.INSURED_CREATED,
            f"Insured processed: {insured_guid}"
        )
    
    def _process_submission(self, transaction: Transaction) -> None:
        """Process the submission step of the workflow"""
        transaction.ims_processing.add_log("Creating submission")
        
        # Extract submission data
        submission_data = self._extract_submission_data(transaction)
        
        # Add insured GUID
        submission_data["insured_guid"] = transaction.ims_processing.insured.guid
        
        # Get producer information
        producer_guid = self._get_producer_guid(transaction)
        submission_data["producer_contact_guid"] = producer_guid
        submission_data["producer_location_guid"] = producer_guid
        
        # Get underwriter - first try by name, then from producer, then use default
        underwriter_guid = self._get_underwriter_guid(transaction)
        if not underwriter_guid:
            underwriter_guid = self.producer_service.get_producer_underwriter(producer_guid)
        
        if not underwriter_guid:
            # Use default underwriter for source system
            source = transaction.parsed_data.get("source_system", "").lower()
            source_config = self._get_source_config(source)
            underwriter_guid = source_config.get("default_underwriter_guid", "00000000-0000-0000-0000-000000000000")
            transaction.ims_processing.add_log(f"Using default underwriter for {source}: {underwriter_guid}")
        else:
            transaction.ims_processing.add_log(f"Using underwriter: {underwriter_guid}")
        
        submission_data["underwriter_guid"] = underwriter_guid
        
        # Create submission
        submission_guid = self.quote_service.create_submission(submission_data)
        transaction.ims_processing.add_log(f"Created submission: {submission_guid}")
        
        # Store submission information
        transaction.ims_processing.submission = IMSSubmission(
            guid=submission_guid,
            insured_guid=submission_data["insured_guid"],
            submission_date=submission_data["submission_date"],
            producer_contact_guid=producer_guid,
            underwriter_guid=underwriter_guid,
            producer_location_guid=producer_guid,
            created_at=datetime.now()
        )
        
        # Update status
        transaction.ims_processing.update_status(
            IMSProcessingStatus.SUBMISSION_CREATED,
            f"Submission created: {submission_guid}"
        )
    
    def _process_quote(self, transaction: Transaction) -> None:
        """Process the quote step of the workflow"""
        transaction.ims_processing.add_log("Creating quote")
        
        # Extract quote data
        quote_data = self._extract_quote_data(transaction)
        
        # Add submission GUID
        quote_data["submission_guid"] = transaction.ims_processing.submission.guid
        
        # Get line of business GUID
        source = transaction.parsed_data.get("source_system", "").lower()
        coverage_type = self._determine_coverage_type(transaction)
        line_guid = self.quote_service.get_default_line_guid(source, coverage_type)
        quote_data["line_guid"] = line_guid
        
        # Set location GUIDs from source configuration
        source = transaction.parsed_data.get("source_system", "").lower()
        source_config = self._get_source_config(source)
        quote_data["quoting_location_guid"] = source_config.get("quoting_location_guid", "00000000-0000-0000-0000-000000000000")
        quote_data["issuing_location_guid"] = source_config.get("issuing_location_guid", "00000000-0000-0000-0000-000000000000")
        quote_data["company_location_guid"] = source_config.get("company_location_guid", "00000000-0000-0000-0000-000000000000")
        
        # Use producer from submission
        quote_data["producer_contact_guid"] = transaction.ims_processing.submission.producer_contact_guid
        
        # Create quote
        quote_guid = self.quote_service.create_quote(quote_data)
        transaction.ims_processing.add_log(f"Created quote: {quote_guid}")
        
        # Store quote information
        transaction.ims_processing.quote = IMSQuote(
            guid=quote_guid,
            submission_guid=quote_data["submission_guid"],
            effective_date=quote_data["effective_date"],
            expiration_date=quote_data["expiration_date"],
            state=quote_data["state"],
            line_guid=line_guid,
            status_id=quote_data.get("status_id", 1),
            created_at=datetime.now()
        )
        
        # Update status
        transaction.ims_processing.update_status(
            IMSProcessingStatus.QUOTE_CREATED,
            f"Quote created: {quote_guid}"
        )
    
    def _process_rating(self, transaction: Transaction) -> None:
        """Process the rating step of the workflow"""
        transaction.ims_processing.add_log("Processing rating")
        
        # Determine rating method
        if self._should_use_excel_rater(transaction):
            self._process_excel_rating(transaction)
        else:
            self._process_manual_rating(transaction)
        
        # Store program-specific data
        self._store_program_data(transaction)
        
        # Update status
        transaction.ims_processing.update_status(
            IMSProcessingStatus.RATED,
            "Quote rated successfully"
        )
    
    def _process_manual_rating(self, transaction: Transaction) -> None:
        """Process manual rating with direct premium pass-through"""
        transaction.ims_processing.add_log("Using manual rating (direct premium pass-through)")
        
        # Extract premium from source system
        premium_data = self._extract_premium_data(transaction)
        source_premium = premium_data.get("total_premium", 0.0)
        
        transaction.ims_processing.add_log(f"Source premium: {source_premium}")
        
        # Create quote option
        option_id = self.quote_service.add_quote_option(transaction.ims_processing.quote.guid)
        transaction.ims_processing.quote.quote_option_id = option_id
        transaction.ims_processing.add_log(f"Created quote option: {option_id}")
        
        # Add premium
        self.quote_service.add_premium(
            transaction.ims_processing.quote.guid,
            option_id,
            source_premium,
            "Premium from source system"
        )
        
        # Update quote premium
        transaction.ims_processing.quote.premium = source_premium
        transaction.ims_processing.add_log(f"Added premium: {source_premium}")
    
    def _process_excel_rating(self, transaction: Transaction) -> None:
        """Process Excel-based rating"""
        transaction.ims_processing.add_log("Using Excel rater")
        
        # Get Excel template
        template_path = self._get_excel_template(transaction)
        if not template_path:
            raise ValueError("Excel template not found")
        
        # Populate template
        populated_path = self._populate_excel_template(template_path, transaction)
        transaction.ims_processing.add_log(f"Populated Excel template: {populated_path}")
        
        # Get rater info
        source = transaction.parsed_data.get("source_system", "").lower()
        coverage_type = self._determine_coverage_type(transaction)
        rater_id, factor_set_guid = self.quote_service.get_rater_info(source, coverage_type)
        
        # Import Excel rater
        result = self.quote_service.import_excel_rater(
            transaction.ims_processing.quote.guid,
            populated_path,
            rater_id,
            factor_set_guid
        )
        
        if result['success'] and result.get('premiums'):
            premium_info = result['premiums'][0]
            transaction.ims_processing.quote.premium = premium_info.get('premium_total', 0.0)
            transaction.ims_processing.quote.quote_option_id = premium_info.get('quote_option_id')
            transaction.ims_processing.add_log(
                f"Excel rating complete. Premium: {transaction.ims_processing.quote.premium}"
            )
        else:
            raise ValueError(f"Excel rating failed: {result.get('error_message', 'Unknown error')}")
    
    def _bind_policy(self, transaction: Transaction) -> None:
        """Bind the policy"""
        transaction.ims_processing.add_log("Binding policy")
        
        if not transaction.ims_processing.quote.quote_option_id:
            raise ValueError("Cannot bind: No quote option ID")
        
        # Bind the quote
        policy_number = self.quote_service.bind_quote(
            transaction.ims_processing.quote.quote_option_id
        )
        transaction.ims_processing.add_log(f"Policy bound: {policy_number}")
        
        # Store policy information
        transaction.ims_processing.policy = IMSPolicy(
            quote_guid=transaction.ims_processing.quote.guid,
            policy_number=policy_number,
            bound_date=self._extract_bound_date(transaction),
            created_at=datetime.now()
        )
        
        # Update status
        transaction.ims_processing.update_status(
            IMSProcessingStatus.BOUND,
            f"Policy bound: {policy_number}"
        )
    
    def _issue_policy(self, transaction: Transaction) -> None:
        """Issue the policy"""
        transaction.ims_processing.add_log("Issuing policy")
        
        policy_number = transaction.ims_processing.policy.policy_number
        
        # Issue the policy
        if self.quote_service.issue_policy(policy_number):
            transaction.ims_processing.add_log(f"Policy issued: {policy_number}")
            
            # Link to external system
            self._link_external_system(transaction)
            
            # Update status
            transaction.ims_processing.update_status(
                IMSProcessingStatus.ISSUED,
                f"Policy issued: {policy_number}"
            )
        else:
            raise ValueError(f"Failed to issue policy: {policy_number}")
    
    def _store_program_data(self, transaction: Transaction) -> None:
        """Store program-specific data"""
        try:
            if not transaction.parsed_data:
                return
            
            source = transaction.parsed_data.get("source_system", "").lower()
            external_id = transaction.parsed_data.get("external_id") or transaction.external_id
            
            if source and transaction.ims_processing.quote:
                self.data_access_service.store_program_data(
                    source,
                    transaction.ims_processing.quote.guid,
                    external_id,
                    transaction.parsed_data
                )
                transaction.ims_processing.add_log(f"Stored {source} program data")
                
        except Exception as e:
            transaction.ims_processing.add_log(f"Warning: Could not store program data: {str(e)}")
    
    def _link_external_system(self, transaction: Transaction) -> None:
        """Link IMS policy to external system"""
        try:
            external_id = transaction.parsed_data.get("external_id") or transaction.external_id
            source = transaction.parsed_data.get("source_system", "Unknown")
            
            if external_id and transaction.ims_processing.quote:
                self.quote_service.update_external_quote_id(
                    transaction.ims_processing.quote.guid,
                    external_id,
                    source
                )
                transaction.ims_processing.add_log(f"Linked to {source}: {external_id}")
                
        except Exception as e:
            transaction.ims_processing.add_log(f"Warning: Could not link external system: {str(e)}")
    
    # Helper methods for data extraction
    
    def _extract_insured_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract insured data from transaction using field mappings"""
        if not transaction.parsed_data:
            raise ValueError("No parsed data available")
        
        data = transaction.parsed_data
        source = data.get("source_system", "triton").lower()
        
        # Get field mapper
        mapper = get_mapper(source)
        
        # Get IMS standard fields
        ims_fields = mapper.get_ims_fields(data)
        
        # Map to our internal structure
        insured_data = {
            "name": ims_fields.get("InsuredName") or data.get("insured_name", "Unknown Insured"),
            "tax_id": data.get("tax_id"),
            "business_type": data.get("business_type"),
            "business_type_id": ims_fields.get("BusinessTypeID"),
            "address": data.get("address") or data.get("insured_address"),
            "city": data.get("city") or data.get("insured_city"),
            "state": ims_fields.get("State") or data.get("state") or data.get("insured_state"),
            "zip_code": ims_fields.get("ZipCode") or data.get("zip_code") or data.get("insured_zip"),
            "source": source  # Pass source for configuration lookup
        }
        
        # Remove None values
        return {k: v for k, v in insured_data.items() if v is not None}
    
    def _extract_location_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract location data from transaction"""
        data = transaction.parsed_data
        
        # Check for locations array
        if "locations" in data and isinstance(data["locations"], list) and data["locations"]:
            return data["locations"][0]
        
        # Extract from flat structure
        return {
            "address": data.get("location_address") or data.get("address", ""),
            "city": data.get("location_city") or data.get("city", ""),
            "state": data.get("location_state") or data.get("state", ""),
            "zip_code": data.get("location_zip") or data.get("zip_code", ""),
            "description": data.get("location_description", "Primary Location")
        }
    
    def _extract_submission_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract submission data from transaction"""
        data = transaction.parsed_data
        
        # Default to today if no date provided
        submission_date = datetime.now().date()
        if data.get("bound_date"):
            try:
                submission_date = datetime.strptime(data["bound_date"], "%Y-%m-%d").date()
            except ValueError:
                pass
        
        return {
            "submission_date": submission_date,
            "producer_contact_guid": "00000000-0000-0000-0000-000000000000",
            "underwriter_guid": "00000000-0000-0000-0000-000000000000",
            "producer_location_guid": "00000000-0000-0000-0000-000000000000"
        }
    
    def _extract_quote_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract quote data from transaction"""
        data = transaction.parsed_data
        
        # Parse dates
        effective_date = datetime.now().date()
        if data.get("effective_date"):
            try:
                effective_date = datetime.strptime(data["effective_date"], "%Y-%m-%d").date()
            except ValueError:
                pass
        
        expiration_date = effective_date.replace(year=effective_date.year + 1)
        if data.get("expiration_date"):
            try:
                expiration_date = datetime.strptime(data["expiration_date"], "%Y-%m-%d").date()
            except ValueError:
                pass
        
        return {
            "effective_date": effective_date,
            "expiration_date": expiration_date,
            "state": data.get("state", "TX"),
            "status_id": 1,  # New
            "billing_type_id": 1  # Agency Bill
        }
    
    def _extract_premium_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract premium data from transaction"""
        data = transaction.parsed_data
        
        total_premium = 0.0
        if data.get("premium"):
            try:
                total_premium = float(data["premium"])
            except (ValueError, TypeError):
                pass
        
        return {"total_premium": total_premium}
    
    def _extract_bound_date(self, transaction: Transaction) -> date:
        """Extract bound date from transaction"""
        data = transaction.parsed_data or {}
        
        if data.get("bound_date"):
            try:
                return datetime.strptime(data["bound_date"], "%Y-%m-%d").date()
            except ValueError:
                pass
        
        return datetime.now().date()
    
    # Helper methods
    
    def _has_location_data(self, transaction: Transaction) -> bool:
        """Check if transaction has location data"""
        data = transaction.parsed_data
        
        if "locations" in data and isinstance(data["locations"], list) and data["locations"]:
            return True
        
        return bool(data.get("location_address") or data.get("address"))
    
    def _get_producer_guid(self, transaction: Transaction) -> str:
        """Get producer GUID for transaction"""
        data = transaction.parsed_data
        
        # Check multiple fields for producer name
        producer_name = (
            data.get("producer_name") or
            (data.get("producer", {}) if isinstance(data.get("producer"), dict) else {}).get("name")
        )
        
        if producer_name:
            transaction.ims_processing.add_log(f"Searching for producer: {producer_name}")
            producer_guid = self.producer_service.get_producer_by_name(producer_name)
            if producer_guid:
                transaction.ims_processing.add_log(f"Found producer: {producer_guid}")
                return producer_guid
            else:
                transaction.ims_processing.add_log(f"Producer not found: {producer_name}")
        
        # Use default for source system
        source = data.get("source_system", "").lower()
        default_guid = self.producer_service.get_default_producer_guid(source)
        transaction.ims_processing.add_log(f"Using default producer for {source}: {default_guid}")
        return default_guid
    
    def _get_underwriter_guid(self, transaction: Transaction) -> Optional[str]:
        """Get underwriter GUID by name"""
        data = transaction.parsed_data
        
        # Check for underwriter name
        underwriter_name = (
            data.get("underwriter_name") or
            (data.get("underwriter", {}) if isinstance(data.get("underwriter"), dict) else {}).get("name")
        )
        
        if underwriter_name:
            transaction.ims_processing.add_log(f"Searching for underwriter: {underwriter_name}")
            underwriter_guid = self.producer_service.find_underwriter_by_name(underwriter_name)
            if underwriter_guid:
                transaction.ims_processing.add_log(f"Found underwriter: {underwriter_guid}")
                return underwriter_guid
            else:
                transaction.ims_processing.add_log(f"Underwriter not found: {underwriter_name}")
        
        return None
    
    def _determine_coverage_type(self, transaction: Transaction) -> str:
        """Determine coverage type from transaction"""
        data = transaction.parsed_data
        
        coverage_type = data.get("coverage_type", "").lower()
        if "excess" in coverage_type:
            return "excess"
        
        # Check line of business
        lob = data.get("line_of_business", "").lower()
        if "excess" in lob:
            return "excess"
        
        return "primary"
    
    def _should_use_excel_rater(self, transaction: Transaction) -> bool:
        """Check if Excel rater should be used"""
        data = transaction.parsed_data or {}
        
        # Only use if explicitly requested
        use_excel = data.get("use_excel_rater", False)
        
        if use_excel:
            # Check if template exists
            template = self._get_excel_template(transaction)
            return template is not None
        
        return False
    
    def _get_excel_template(self, transaction: Transaction) -> Optional[str]:
        """Get Excel template path for transaction"""
        data = transaction.parsed_data or {}
        
        # Generate template paths to check
        source = data.get("source_system", "").lower()
        lob = data.get("line_of_business", "").lower().replace(" ", "_")
        state = data.get("state", "").upper()
        
        templates = [
            f"{source}_{lob}_{state}.xlsx",
            f"{source}_{lob}.xlsx",
            f"{source}.xlsx",
            "default.xlsx"
        ]
        
        for template in templates:
            path = os.path.join(self.templates_dir, template)
            if os.path.exists(path):
                return path
        
        return None
    
    def _populate_excel_template(self, template_path: str, transaction: Transaction) -> str:
        """Populate Excel template with transaction data using field mappings"""
        if not transaction.parsed_data:
            raise ValueError("Cannot populate Excel template: Missing parsed data")
        
        data = transaction.parsed_data
        source = data.get("source_system", "triton").lower()
        
        # Get field mapper
        mapper = get_mapper(source)
        
        # Get Excel fields
        excel_fields = mapper.get_excel_fields(data)
        
        # Create a copy of the template in a temp file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_path = temp_file.name
        
        try:
            # Import openpyxl here to avoid import errors if not installed
            import openpyxl
            
            # Load the template
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Find headers in the first row
            headers = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers[header] = col
            
            # Populate data in row 2
            row = 2
            for field_name, value in excel_fields.items():
                if field_name in headers:
                    ws.cell(row=row, column=headers[field_name], value=value)
                    transaction.ims_processing.add_log(f"Set Excel field {field_name} = {value}")
            
            # Add any additional standard fields
            if "Policy_Number" in headers and data.get("policy_number"):
                ws.cell(row=row, column=headers["Policy_Number"], value=data["policy_number"])
            
            # Save the populated workbook
            wb.save(temp_path)
            
            transaction.ims_processing.add_log(f"Populated Excel template saved to: {temp_path}")
            return temp_path
            
        except ImportError:
            logger.error("openpyxl not installed - cannot populate Excel template")
            # Return original template as fallback
            return template_path
        except Exception as e:
            logger.error(f"Error populating Excel template: {str(e)}")
            # Return original template as fallback
            return template_path
    
    def _has_additional_insureds(self, transaction: Transaction) -> bool:
        """Check if transaction has additional insureds"""
        data = transaction.parsed_data
        
        # Check for additional_insured field
        if "additional_insured" in data:
            additional = data["additional_insured"]
            if isinstance(additional, list) and additional:
                return True
            elif isinstance(additional, dict) and (additional.get("name") or additional.get("address")):
                return True
        
        # Check for additional_insureds (plural)
        if "additional_insureds" in data and isinstance(data["additional_insureds"], list):
            return len(data["additional_insureds"]) > 0
        
        return False
    
    def _process_additional_insureds(self, transaction: Transaction, primary_insured_guid: str) -> None:
        """Process additional insureds and link them to the primary insured"""
        data = transaction.parsed_data
        additional_insureds = []
        
        # Normalize to list
        if "additional_insured" in data:
            additional = data["additional_insured"]
            if isinstance(additional, dict):
                additional_insureds = [additional]
            elif isinstance(additional, list):
                additional_insureds = additional
        elif "additional_insureds" in data:
            additional_insureds = data["additional_insureds"]
        
        transaction.ims_processing.add_log(f"Processing {len(additional_insureds)} additional insureds")
        
        for idx, additional in enumerate(additional_insureds):
            try:
                # Extract additional insured data
                ai_name = additional.get("name", f"Additional Insured {idx + 1}")
                ai_address = additional.get("address", "")
                
                if not ai_name and not ai_address:
                    continue
                
                transaction.ims_processing.add_log(f"Processing additional insured: {ai_name}")
                
                # Create search criteria
                search_criteria = {
                    "name": ai_name,
                    "address": ai_address
                }
                
                # Try to find existing additional insured
                match = self.insured_service.matcher.find_best_match(search_criteria)
                
                if match:
                    ai_guid = match.get("InsuredGUID")
                    transaction.ims_processing.add_log(f"Found existing additional insured: {ai_guid}")
                else:
                    # Create new additional insured
                    ai_data = {
                        "name": ai_name,
                        "business_type_id": 6,  # Other
                        "address": ai_address
                    }
                    ai_guid = self.insured_service.create_insured(ai_data)
                    transaction.ims_processing.add_log(f"Created additional insured: {ai_guid}")
                
                # Link additional insured to primary insured
                # This would typically be done through a custom stored procedure
                self._link_additional_insured(primary_insured_guid, ai_guid, transaction)
                
            except Exception as e:
                transaction.ims_processing.add_log(f"Error processing additional insured {idx + 1}: {str(e)}")
    
    def _link_additional_insured(self, primary_guid: str, additional_guid: str, transaction: Transaction) -> None:
        """Link an additional insured to the primary insured"""
        try:
            # This would call a custom stored procedure to create the relationship
            params = {
                "PrimaryInsuredGUID": primary_guid,
                "AdditionalInsuredGUID": additional_guid,
                "QuoteGUID": str(transaction.ims_processing.quote.guid) if transaction.ims_processing.quote else ""
            }
            
            # Log the linkage for now
            transaction.ims_processing.add_log(
                f"Linked additional insured {additional_guid} to primary {primary_guid}"
            )
            
            # In production, this would call:
            # self.data_access_service.execute_command("LinkAdditionalInsured", params)
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error linking additional insured: {str(e)}")
    
    def _handle_workflow_error(self, transaction: Transaction, error: Exception) -> None:
        """Handle workflow errors"""
        import traceback
        error_details = traceback.format_exc()
        
        logger.error(f"Workflow error for transaction {transaction.transaction_id}: {str(error)}")
        logger.error(f"Stack trace:\n{error_details}")
        
        transaction.ims_processing.add_log(f"ERROR: {str(error)}")
        transaction.ims_processing.add_log(f"Stage: {transaction.ims_processing.status.value}")
        
        transaction.update_status(
            TransactionStatus.FAILED,
            f"IMS workflow error at {transaction.ims_processing.status.value}: {str(error)}"
        )
        transaction.ims_processing.status = IMSProcessingStatus.ERROR