import logging
import os
import base64
from datetime import datetime, date
from typing import Dict, Any, Optional, Union, Tuple, List
import uuid
import json
import tempfile
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from app.models.transaction_models import (
    Transaction, TransactionStatus, IMSProcessingStatus,
    IMSInsured, IMSSubmission, IMSQuote, IMSPolicy
)
from app.services.ims_soap_client import IMSSoapClient
from app.core.config import settings

logger = logging.getLogger(__name__)

class IMSWorkflowService:
    """
    Handles the workflow for processing transactions through IMS
    """
    
    def __init__(self, environment=None):
        """Initialize with environment settings"""
        env = environment or settings.DEFAULT_ENVIRONMENT
        env_config = settings.IMS_ENVIRONMENTS.get(env)
        if not env_config:
            raise ValueError(f"Unknown environment: {env}")
        
        self.config_file = env_config["config_file"]
        self.username = env_config["username"]
        self.password = env_config["password"]
        self.env_config = env_config
        self.soap_client = IMSSoapClient(self.config_file, env_config)
        self.token = None
        
        # Templates directory for Excel raters
        self.templates_dir = os.path.join(os.path.dirname(__file__), "../../templates")
        os.makedirs(self.templates_dir, exist_ok=True)
        
        # Cached lookup data
        self.business_types = None
        self.line_guids = None
        self.factor_sets = None
        self.rater_ids = None
    
    def authenticate(self) -> str:
        """Authenticate with IMS and get a token"""
        logger.info(f"Authenticating with IMS using {self.username}")
        self.token = self.soap_client.login(self.username, self.password)
        return self.token
    
    def process_transaction(self, transaction: Transaction) -> Transaction:
        """
        Process a transaction with IMS integration
        """
        logger.info(f"Processing transaction {transaction.transaction_id} in IMS")
        
        try:
            # Ensure we're authenticated
            if not self.token:
                self.authenticate()
            
            # Process based on transaction state
            if transaction.ims_processing.status == IMSProcessingStatus.PENDING:
                self._process_insured(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.INSURED_CREATED:
                self._process_submission(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.SUBMISSION_CREATED:
                self._process_quote(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.QUOTE_CREATED:
                self._process_rating(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.RATED:
                self._bind_policy(transaction)
                
            if transaction.ims_processing.status == IMSProcessingStatus.BOUND:
                self._issue_policy(transaction)
                
            # If we got here without errors, update the transaction status
            if transaction.ims_processing.status == IMSProcessingStatus.ISSUED:
                transaction.update_status(
                    TransactionStatus.COMPLETED,
                    f"Transaction successfully processed in IMS. Policy number: {transaction.ims_processing.policy.policy_number}"
                )
                
            return transaction
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"Error processing transaction {transaction.transaction_id} in IMS: {str(e)}")
            logger.error(f"Full traceback:\n{error_details}")
            
            # Add detailed error to processing logs
            transaction.ims_processing.add_log(f"ERROR: {str(e)}")
            transaction.ims_processing.add_log(f"Error occurred at stage: {transaction.ims_processing.status.value}")
            transaction.ims_processing.add_log(f"Stack trace: {error_details}")
            
            transaction.update_status(
                TransactionStatus.FAILED,
                f"Error processing transaction in IMS at stage {transaction.ims_processing.status.value}: {str(e)}"
            )
            transaction.ims_processing.status = IMSProcessingStatus.ERROR
            return transaction
    
    def _process_insured(self, transaction: Transaction) -> None:
        """Process the insured data for the transaction"""
        transaction.ims_processing.add_log("Processing insured data")
        
        # Ensure we're authenticated
        if not self.soap_client.token:
            transaction.ims_processing.add_log(f"Authenticating with IMS using {self.username}")
            try:
                self.authenticate()
                transaction.ims_processing.add_log(f"Authentication successful")
            except Exception as e:
                transaction.ims_processing.add_log(f"Authentication failed: {str(e)}")
                raise
        
        # Extract insured data from transaction
        insured_data = self._extract_insured_data(transaction)
        
        # Check if insured exists
        try:
            transaction.ims_processing.add_log(f"Looking for existing insured: {insured_data['name']}")
            insured_guid = self.soap_client.find_insured_by_name(
                insured_data["name"], 
                insured_data.get("tax_id"),
                city=insured_data.get("city", ""),
                state=insured_data.get("state", ""),
                zip_code=insured_data.get("zip_code", "")
            )
            
            if insured_guid:
                transaction.ims_processing.add_log(f"Found existing insured with GUID: {insured_guid}")
            else:
                # Create new insured
                transaction.ims_processing.add_log(f"Creating new insured: {insured_data['name']}")
                insured_guid = self.soap_client.add_insured(insured_data)
                transaction.ims_processing.add_log(f"Created insured with GUID: {insured_guid}")
                
                # Add insured location if provided
                if transaction.parsed_data and "locations" in transaction.parsed_data and \
                   isinstance(transaction.parsed_data["locations"], list) and \
                   len(transaction.parsed_data["locations"]) > 0:
                    
                    location_data = transaction.parsed_data["locations"][0]
                    transaction.ims_processing.add_log(f"Adding location for insured: {location_data.get('address', '')}")
                    location_id = self.soap_client.add_insured_location(insured_guid, location_data)
                    transaction.ims_processing.add_log(f"Added location with ID: {location_id}")
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error processing insured: {str(e)}")
            raise
        
        # Store insured information
        transaction.ims_processing.insured = IMSInsured(
            guid=insured_guid,
            name=insured_data["name"],
            tax_id=insured_data.get("tax_id"),
            business_type_id=insured_data.get("business_type_id"),
            created_at=datetime.now()
        )
        
        # Update status
        transaction.ims_processing.update_status(
            IMSProcessingStatus.INSURED_CREATED,
            f"Insured processed successfully: {insured_guid}"
        )
    
    def _process_submission(self, transaction: Transaction) -> None:
        """Process the submission for the transaction"""
        transaction.ims_processing.add_log("Creating submission")
        
        # Extract submission data
        submission_data = self._extract_submission_data(transaction)
        
        # Add insured GUID from previous step
        submission_data["insured_guid"] = transaction.ims_processing.insured.guid
        
        try:
            # Create submission
            transaction.ims_processing.add_log(f"Creating submission for insured: {submission_data['insured_guid']}")
            submission_guid = self.soap_client.add_submission(submission_data)
            transaction.ims_processing.add_log(f"Created submission with GUID: {submission_guid}")
            
            # Store submission information
            transaction.ims_processing.submission = IMSSubmission(
                guid=submission_guid,
                insured_guid=submission_data["insured_guid"],
                submission_date=submission_data["submission_date"],
                producer_contact_guid=submission_data.get("producer_contact_guid"),
                underwriter_guid=submission_data.get("underwriter_guid"),
                producer_location_guid=submission_data.get("producer_location_guid"),
                created_at=datetime.now()
            )
            
            # Update status
            transaction.ims_processing.update_status(
                IMSProcessingStatus.SUBMISSION_CREATED,
                f"Submission created successfully: {submission_guid}"
            )
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error creating submission: {str(e)}")
            raise
    
    def _process_quote(self, transaction: Transaction) -> None:
        """Process the quote for the transaction"""
        transaction.ims_processing.add_log("Creating quote")
        
        # Extract quote data
        quote_data = self._extract_quote_data(transaction)
        
        # Add submission GUID from previous step
        quote_data["submission_guid"] = transaction.ims_processing.submission.guid
        
        try:
            # Create quote
            transaction.ims_processing.add_log(f"Creating quote for submission: {quote_data['submission_guid']}")
            quote_guid = self.soap_client.add_quote(quote_data)
            transaction.ims_processing.add_log(f"Created quote with GUID: {quote_guid}")
            
            # Store quote information
            transaction.ims_processing.quote = IMSQuote(
                guid=quote_guid,
                submission_guid=quote_data["submission_guid"],
                effective_date=quote_data["effective_date"],
                expiration_date=quote_data["expiration_date"],
                state=quote_data["state"],
                line_guid=quote_data.get("line_guid"),
                status_id=quote_data.get("status_id"),
                created_at=datetime.now()
            )
            
            # Update status
            transaction.ims_processing.update_status(
                IMSProcessingStatus.QUOTE_CREATED,
                f"Quote created successfully: {quote_guid}"
            )
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error creating quote: {str(e)}")
            raise
    
    def _process_rating(self, transaction: Transaction) -> None:
        """Process the rating for the transaction"""
        transaction.ims_processing.add_log("Processing rating data")
        
        # Determine which rating method to use
        if self._should_use_excel_rater(transaction):
            self._process_excel_rating(transaction)
        else:
            self._process_manual_rating(transaction)
        
        # Update status
        transaction.ims_processing.update_status(
            IMSProcessingStatus.RATED,
            "Quote rated successfully"
        )
    
    def _process_excel_rating(self, transaction: Transaction) -> None:
        """Process rating using Excel rater"""
        transaction.ims_processing.add_log("Using Excel rater")
        
        try:
            # Get the Excel template and populate it
            template_path = self._get_excel_template(transaction)
            if not template_path:
                raise ValueError("Excel template not found for this product/program")
            
            transaction.ims_processing.add_log(f"Using Excel template: {template_path}")
            
            # Create a populated Excel file
            populated_file_path = self._populate_excel_template(template_path, transaction)
            transaction.ims_processing.add_log(f"Populated Excel file created at: {populated_file_path}")
            
            # Store the path for future reference
            transaction.ims_processing.excel_rater_file_path = populated_file_path
            
            # Upload the Excel file to IMS
            transaction.ims_processing.add_log(f"Uploading Excel file to IMS for quote: {transaction.ims_processing.quote.guid}")
            rating_result = self._import_excel_rater(
                transaction.ims_processing.quote.guid,
                populated_file_path,
                transaction.ims_processing.excel_rater_id,
                transaction.ims_processing.factor_set_guid
            )
            
            # Store the rating results
            if transaction.ims_processing.quote:
                transaction.ims_processing.quote.premium = rating_result.get("premium_total")
                transaction.ims_processing.quote.quote_option_id = rating_result.get("quote_option_id")
            
            transaction.ims_processing.add_log(
                f"Excel rating completed. Premium: {rating_result.get('premium_total')}, "
                f"QuoteOptionID: {rating_result.get('quote_option_id')}"
            )
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error processing Excel rating: {str(e)}")
            raise
    
    def _process_manual_rating(self, transaction: Transaction) -> None:
        """Process rating manually with direct premium pass-through from source system"""
        transaction.ims_processing.add_log("Using manual rating (direct premium pass-through)")
        
        try:
            # Extract premium data from transaction (NO CALCULATION - direct pass-through)
            premium_data = self._extract_premium_data(transaction)
            source_premium = premium_data.get("total_premium", 0.0)
            
            transaction.ims_processing.add_log(f"Source system premium: {source_premium} (direct pass-through)")
            
            # Create quote option if needed
            if not transaction.ims_processing.quote.quote_option_id:
                transaction.ims_processing.add_log(f"Creating quote option for quote: {transaction.ims_processing.quote.guid}")
                option_id = self.soap_client.add_quote_option(transaction.ims_processing.quote.guid)
                transaction.ims_processing.quote.quote_option_id = option_id
                transaction.ims_processing.add_log(f"Created quote option: {option_id}")
            
            # Add premium to quote (direct from source system - no pro-rata calculation)
            transaction.ims_processing.add_log(f"Adding premium: {source_premium} (direct from source system)")
            self.soap_client.add_premium(
                transaction.ims_processing.quote.guid,
                transaction.ims_processing.quote.quote_option_id,
                source_premium,  # Direct pass-through - NO CALCULATION
                "Premium from source system (direct pass-through)"
            )
            
            # Store program-specific data if available
            self._store_program_specific_data(transaction)
            
            # Update premium in quote
            transaction.ims_processing.quote.premium = source_premium
            transaction.ims_processing.add_log(
                f"Manual rating completed. Premium: {source_premium} (direct pass-through from source)"
            )
        
        except Exception as e:
            transaction.ims_processing.add_log(f"Error processing manual rating: {str(e)}")
            raise
    
    def _bind_policy(self, transaction: Transaction) -> None:
        """Bind the policy for the transaction"""
        transaction.ims_processing.add_log("Binding policy")
        
        if not transaction.ims_processing.quote.quote_option_id:
            raise ValueError("Cannot bind policy: Quote option ID not found")
        
        try:
            # Bind the policy
            transaction.ims_processing.add_log(f"Binding quote option: {transaction.ims_processing.quote.quote_option_id}")
            policy_number = self.soap_client.bind(transaction.ims_processing.quote.quote_option_id)
            transaction.ims_processing.add_log(f"Policy bound: {policy_number}")
            
            # Store policy information
            bound_date = datetime.now().date()
            if transaction.parsed_data and isinstance(transaction.parsed_data, dict):
                if transaction.parsed_data.get("bound_date"):
                    try:
                        bound_date = datetime.strptime(
                            transaction.parsed_data["bound_date"], "%Y-%m-%d"
                        ).date()
                    except ValueError:
                        pass
            
            transaction.ims_processing.policy = IMSPolicy(
                quote_guid=transaction.ims_processing.quote.guid,
                policy_number=policy_number,
                bound_date=bound_date,
                created_at=datetime.now()
            )
            
            # Update status
            transaction.ims_processing.update_status(
                IMSProcessingStatus.BOUND,
                f"Policy bound successfully: {policy_number}"
            )
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error binding policy: {str(e)}")
            raise
    
    def _issue_policy(self, transaction: Transaction) -> None:
        """Issue the policy for the transaction"""
        transaction.ims_processing.add_log("Issuing policy")
        
        try:
            # Issue the policy
            if transaction.ims_processing.policy and transaction.ims_processing.policy.policy_number:
                policy_number = transaction.ims_processing.policy.policy_number
                transaction.ims_processing.add_log(f"Issuing policy: {policy_number}")
                result = self.soap_client.issue_policy(policy_number)
                
                if result:
                    transaction.ims_processing.add_log(f"Policy issued successfully: {policy_number}")
                    
                    # Link back to source system using UpdateExternalQuoteId
                    self._link_external_system(transaction)
                    
                    # Update status
                    transaction.ims_processing.update_status(
                        IMSProcessingStatus.ISSUED,
                        f"Policy issued successfully: {policy_number}"
                    )
                else:
                    transaction.ims_processing.add_log(f"Policy issue failed: {policy_number}")
                    raise ValueError(f"Failed to issue policy: {policy_number}")
                    
        except Exception as e:
            transaction.ims_processing.add_log(f"Error issuing policy: {str(e)}")
            raise
    
    def _store_program_specific_data(self, transaction: Transaction) -> None:
        """Store program-specific data in custom tables using stored procedures"""
        try:
            if not transaction.parsed_data:
                return
                
            # Determine program type from transaction data
            program = transaction.parsed_data.get("program", "").lower()
            source_system = transaction.parsed_data.get("source_system", "").lower()
            
            # Store data based on program type
            if program == "triton" or source_system == "triton":
                self._store_triton_data(transaction)
            elif program == "xuber" or source_system == "xuber":
                self._store_xuber_data(transaction)
            else:
                # Store generic program data
                self._store_generic_program_data(transaction)
                
        except Exception as e:
            transaction.ims_processing.add_log(f"Warning: Could not store program-specific data: {str(e)}")
            # Don't raise exception - this is supplementary data storage
    
    def _store_triton_data(self, transaction: Transaction) -> None:
        """Store Triton-specific data in custom tables"""
        try:
            # Call custom stored procedure for Triton data
            external_id = transaction.parsed_data.get("external_id") or transaction.external_id
            
            parameters = [
                ("QuoteGUID", str(transaction.ims_processing.quote.guid)),
                ("ExternalPolicyID", external_id),
                ("OriginalPremium", str(transaction.ims_processing.quote.premium or 0.0)),
                ("SourceSystemData", json.dumps(transaction.parsed_data))
            ]
            
            # Add vehicle and location data if available
            if transaction.parsed_data.get("vehicles"):
                parameters.append(("VehicleData", json.dumps(transaction.parsed_data["vehicles"])))
            if transaction.parsed_data.get("locations"):
                parameters.append(("LocationData", json.dumps(transaction.parsed_data["locations"])))
            
            transaction.ims_processing.add_log("Storing Triton-specific data in custom tables")
            # Note: This would use ExecuteCommand to call StoreRSGTritonData_WS
            # self.soap_client.execute_command("StoreRSGTritonData", parameters)
            
        except Exception as e:
            transaction.ims_processing.add_log(f"Error storing Triton data: {str(e)}")
            # Don't raise - supplementary data storage
    
    def _store_xuber_data(self, transaction: Transaction) -> None:
        """Store Xuber-specific data in custom tables"""
        try:
            # Similar implementation for Xuber program
            transaction.ims_processing.add_log("Storing Xuber-specific data in custom tables")
            # Implementation would be similar to Triton but for Xuber-specific fields
        except Exception as e:
            transaction.ims_processing.add_log(f"Error storing Xuber data: {str(e)}")
    
    def _store_generic_program_data(self, transaction: Transaction) -> None:
        """Store generic program data in custom tables"""
        try:
            transaction.ims_processing.add_log("Storing generic program data")
            # Generic data storage for unknown programs
        except Exception as e:
            transaction.ims_processing.add_log(f"Error storing generic program data: {str(e)}")
    
    def _link_external_system(self, transaction: Transaction) -> None:
        """Link IMS policy back to source system using UpdateExternalQuoteId"""
        try:
            external_id = transaction.parsed_data.get("external_id") or transaction.external_id
            source_system = transaction.parsed_data.get("source_system", "Unknown")
            
            if external_id and transaction.ims_processing.quote:
                transaction.ims_processing.add_log(f"Linking to external system: {external_id}")
                
                # Use UpdateExternalQuoteId for future lookups
                self.soap_client.update_external_quote_id(
                    transaction.ims_processing.quote.guid,
                    external_id,
                    source_system
                )
                
                transaction.ims_processing.add_log(f"Successfully linked to external system {source_system}: {external_id}")
                
        except Exception as e:
            transaction.ims_processing.add_log(f"Warning: Could not link external system: {str(e)}")
            # Don't raise - this is supplementary functionality
    
    def _import_excel_rater(self, quote_guid: str, excel_file_path: str, 
                           rater_id: int, factor_set_guid: str) -> Dict[str, Any]:
        """Import Excel rater to IMS"""
        try:
            logger.info(f"Importing Excel rater for quote: {quote_guid}")
            
            # Read the file and convert to base64
            with open(excel_file_path, "rb") as file:
                file_bytes = base64.b64encode(file.read()).decode("utf-8")
            
            # Call the IMS API to import the Excel rater
            result = self.soap_client.import_excel_rater(
                quote_guid, 
                file_bytes, 
                os.path.basename(excel_file_path), 
                rater_id, 
                factor_set_guid, 
                True
            )
            
            if result['success']:
                logger.info(f"Successfully imported Excel rater")
                
                # Find the first premium option
                if result['premiums'] and len(result['premiums']) > 0:
                    premium = result['premiums'][0]
                    return {
                        "success": True,
                        "premium_total": premium['premium_total'],
                        "fee_total": premium['fee_total'],
                        "quote_option_id": premium['quote_option_id']
                    }
                
                return {
                    "success": True,
                    "premium_total": 0.0,
                    "fee_total": 0.0,
                    "quote_option_id": None
                }
            else:
                logger.error(f"Failed to import Excel rater: {result.get('error_message')}")
                return {
                    "success": False,
                    "error_message": result.get('error_message')
                }
                
        except Exception as e:
            logger.error(f"Error importing Excel rater: {str(e)}")
            raise
    
    # Helper methods for data extraction and transformation
    
    def _extract_insured_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract insured data from transaction"""
        # This method would extract and transform insured data from the transaction
        # to match IMS API requirements
        
        # Use processed_data if available (from transformers), otherwise use parsed_data
        data = transaction.processed_data or transaction.parsed_data
        if not data:
            raise ValueError("Cannot extract insured data: Missing data")
        
        # If we have processed_data with insured_data section, use that structure
        if transaction.processed_data and "insured_data" in transaction.processed_data:
            insured_data = transaction.processed_data["insured_data"]
            transaction.ims_processing.add_log(f"Using transformed insured data: {insured_data.get('name')}, business_type_id={insured_data.get('business_type_id')}")
            return insured_data
        
        # Handle both nested and flat data structures
        # Check if data is nested under 'policy' key (from CSV loader)
        if "policy" in data:
            policy_data = data["policy"]
        else:
            policy_data = data
            
        # Log the data structure for debugging
        transaction.ims_processing.add_log(f"Extracting insured data from structure: {list(policy_data.keys())[:10]}")
        
        # Extract insured name - try multiple possible field names
        insured_name = (
            policy_data.get("insured_name") or 
            policy_data.get("insured", {}).get("name") or
            policy_data.get("insuredName") or
            "Unknown Insured"
        )
        
        # Extract business_type_id from transformed data
        business_type_id = None
        
        # Check for business_type_id in various locations
        if "insured_data" in policy_data:
            business_type_id = policy_data["insured_data"].get("business_type_id")
        elif "business_type_id" in policy_data:
            business_type_id = policy_data["business_type_id"]
        
        # If still not found, default to 13 (Corporation in IMS_DEV)
        if business_type_id is None:
            business_type_id = 13
            transaction.ims_processing.add_log(f"No business_type_id found, defaulting to Corporation (13)")
        else:
            transaction.ims_processing.add_log(f"Using business_type_id: {business_type_id}")
        
        result = {
            "name": insured_name,
            "tax_id": policy_data.get("tax_id") or policy_data.get("insured", {}).get("tax_id"),
            "business_type_id": business_type_id
        }
        
        # Extract contact information - handle flat or nested structure
        result.update({
            "address": policy_data.get("insured_address") or policy_data.get("address"),
            "city": policy_data.get("insured_city") or policy_data.get("city"),
            "state": policy_data.get("insured_state") or policy_data.get("state"),
            "zip_code": policy_data.get("insured_zip") or policy_data.get("zip_code")
        })
        
        # If contact info is nested, also check there
        if "insured" in policy_data and "contact" in policy_data["insured"]:
            contact = policy_data["insured"]["contact"]
            result.update({
                "contact_name": contact.get("name") or result.get("contact_name"),
                "contact_email": contact.get("email") or result.get("contact_email"),
                "contact_phone": contact.get("phone") or result.get("contact_phone"),
                "address": contact.get("address") or result.get("address"),
                "city": contact.get("city") or result.get("city"),
                "state": contact.get("state") or result.get("state"),
                "zip_code": contact.get("zip_code") or result.get("zip_code")
            })
        
        transaction.ims_processing.add_log(f"Extracted insured data: name={result['name']}, address={result.get('address')}")
        
        return result
    
    def _extract_submission_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract submission data from transaction"""
        # This method would extract and transform submission data from the transaction
        # to match IMS API requirements
        
        # Use processed_data if available (from transformers), otherwise use parsed_data
        data = transaction.processed_data or transaction.parsed_data
        if not data:
            raise ValueError("Cannot extract submission data: Missing data")
        
        # Check if we have transformed data with policy_data section
        if transaction.processed_data and "policy_data" in transaction.processed_data:
            # Use the transformed data structure
            policy_data = transaction.processed_data["policy_data"]
            producer_data = transaction.processed_data.get("producer_data", {})
            
            # Get dates from policy_data
            submission_date = policy_data.get("bound_date") or datetime.now().date()
            if isinstance(submission_date, str):
                try:
                    submission_date = datetime.strptime(submission_date, "%Y-%m-%d").date()
                except ValueError:
                    submission_date = datetime.now().date()
        else:
            # Fall back to original logic for non-transformed data
            submission_date = datetime.now().date()
            if data.get("bound_date"):
                try:
                    submission_date = datetime.strptime(data["bound_date"], "%Y-%m-%d").date()
                except ValueError:
                    pass
            policy_data = data
            producer_data = data.get("producer_data", {})
        
        # Get source-specific defaults from configuration
        source = transaction.source or "triton"
        source_config = self.env_config.get("sources", {}).get(source, {})
        
        # Get the default producer location GUID from configuration
        producer_location_guid = source_config.get("default_producer_guid", "00000000-0000-0000-0000-000000000000")
        transaction.ims_processing.add_log(f"Using producer location GUID from configuration: {producer_location_guid}")
        
        # Try to get the producer contact for this location
        producer_contact_guid = None
        
        # Use the new producer search chain
        try:
            # First try to get producer name from the data
            producer_name = None
            
            # Check if we have transformed data
            if transaction.processed_data and "producer_data" in transaction.processed_data:
                producer_name = transaction.processed_data["producer_data"].get("name")
            else:
                # Fall back to original logic
                producer_info = data.get("producer") or data.get("producer_data")
                if producer_info and isinstance(producer_info, dict):
                    producer_name = producer_info.get("name") or producer_info.get("producer_name")
                elif isinstance(data.get("producer_name"), str):
                    producer_name = data.get("producer_name")
            
            if producer_name:
                transaction.ims_processing.add_log(f"Searching for producer: {producer_name}")
                
                # Use ProducerSearch to find the producer
                search_results = self.soap_client.producer_search(producer_name)
                
                if search_results:
                    # Use the first matching result
                    producer_result = search_results[0]
                    location_code = producer_result.get('location_code')
                    
                    if location_code:
                        transaction.ims_processing.add_log(f"Found producer location code: {location_code}")
                        
                        # Get the contact GUID for this location
                        contact_guid = self.soap_client.get_producer_contact_by_location_code(location_code)
                        
                        if contact_guid:
                            producer_contact_guid = contact_guid
                            producer_location_guid = producer_result.get('producer_location_guid')
                            transaction.ims_processing.add_log(f"Found producer contact GUID: {producer_contact_guid}")
                            transaction.ims_processing.add_log(f"Using producer location GUID: {producer_location_guid}")
                        else:
                            transaction.ims_processing.add_log(f"Warning: No contact found for location code: {location_code}")
                    else:
                        transaction.ims_processing.add_log("Warning: No location code in search result")
                else:
                    transaction.ims_processing.add_log(f"Warning: No producers found for name: {producer_name}")
            else:
                transaction.ims_processing.add_log("No producer name provided in data")
        except Exception as e:
            transaction.ims_processing.add_log(f"Error in producer search chain: {str(e)}")
            import traceback
            transaction.ims_processing.add_log(f"Traceback: {traceback.format_exc()}")
        
        # If we still don't have a producer contact, use the default
        if not producer_contact_guid:
            producer_contact_guid = producer_location_guid
            transaction.ims_processing.add_log(f"Using default producer location GUID as contact GUID: {producer_contact_guid}")
        
        # Note: The producer lookup by name happens below in the producer extraction section
        # That will override these defaults if a producer is found
        
        # TODO: Re-enable this once we have valid producer data in IMS
        # try:
        #     # First get producer info to get the location code
        #     transaction.ims_processing.add_log(f"Looking up producer info for GUID: {producer_location_guid}")
        #     producer_info = self.soap_client.get_producer_info(producer_location_guid)
        #     
        #     if producer_info:
        #         transaction.ims_processing.add_log(f"Got producer info: {producer_info.get('producer_name')} - {producer_info.get('location_name')}")
        #         location_code = producer_info.get("location_code")
        #         
        #         if location_code:
        #             transaction.ims_processing.add_log(f"Found producer location code: {location_code}")
        #             
        #             # Now get the contact for this location
        #             producer_contact_guid = self.soap_client.get_producer_contact_by_location_code(location_code)
        #             if producer_contact_guid:
        #                 transaction.ims_processing.add_log(f"Found producer contact GUID: {producer_contact_guid}")
        #             else:
        #                 transaction.ims_processing.add_log("Warning: No producer contact found for location")
        #         else:
        #             transaction.ims_processing.add_log(f"Warning: No location code in producer info: {list(producer_info.keys())}")
        #     else:
        #         transaction.ims_processing.add_log("Warning: get_producer_info returned None")
        # except Exception as e:
        #     transaction.ims_processing.add_log(f"Warning: Error looking up producer contact: {str(e)}")
        #     import traceback
        #     transaction.ims_processing.add_log(f"Traceback: {traceback.format_exc()}")
        
        result = {
            "insured_guid": None,  # Will be filled in by the caller
            "submission_date": submission_date,
            "producer_contact_guid": producer_contact_guid,  # Use contact GUID
            "underwriter_guid": "E4391D2A-58FB-4E2D-8B7D-3447D9E18C88",  # Valid underwriter GUID
            "producer_location_guid": "3ED4DB47-9867-41A7-840D-61A19528F1F1",  # Use correct producer location GUID
        }
        
        # Producer information is already handled in the producer search chain above
        
        # Extract underwriter information if available
        underwriter_name = None
        
        # Check if we have transformed data
        if transaction.processed_data and "producer_data" in transaction.processed_data:
            underwriter_name = transaction.processed_data["producer_data"].get("underwriter_name")
        else:
            # Fall back to original logic
            if producer_data and isinstance(producer_data, dict):
                underwriter_name = producer_data.get("underwriter_name")
            if not underwriter_name:
                underwriter_name = data.get("underwriter") or data.get("underwriter_name")
        
        if underwriter_name:
            try:
                user_info = self.soap_client.get_user_by_name(underwriter_name)
                if user_info and user_info.get("user_guid"):
                    result["underwriter_guid"] = user_info["user_guid"]
                    transaction.ims_processing.add_log(f"Found underwriter: {user_info['first_name']} {user_info['last_name']} ({user_info['user_guid']})")
                else:
                    transaction.ims_processing.add_log(f"Warning: Could not find underwriter: {underwriter_name}")
            except Exception as e:
                transaction.ims_processing.add_log(f"Warning: Error looking up underwriter: {str(e)}")
        
        return result
    
    def _extract_quote_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract quote data from transaction"""
        # This method would extract and transform quote data from the transaction
        # to match IMS API requirements
        
        # Use processed_data if available (from transformers), otherwise use parsed_data
        data = transaction.processed_data or transaction.parsed_data
        if not data:
            raise ValueError("Cannot extract quote data: Missing data")
        
        # Check if we have transformed data
        if transaction.processed_data and "policy_data" in transaction.processed_data:
            # Use the transformed data structure
            policy_data = transaction.processed_data["policy_data"]
            insured_data = transaction.processed_data.get("insured_data", {})
            
            effective_date = policy_data.get("effective_date", datetime.now().date())
            expiration_date = policy_data.get("expiration_date", datetime.now().date().replace(year=datetime.now().year + 1))
            
            # Convert dates if they're strings
            if isinstance(effective_date, str):
                try:
                    effective_date = datetime.strptime(effective_date, "%Y-%m-%d").date()
                except ValueError:
                    effective_date = datetime.now().date()
            if isinstance(expiration_date, str):
                try:
                    expiration_date = datetime.strptime(expiration_date, "%Y-%m-%d").date()
                except ValueError:
                    expiration_date = datetime.now().date().replace(year=datetime.now().year + 1)
            
            state = insured_data.get("state", "TX")
            line_of_business = policy_data.get("line_of_business", "General Liability")
            line_guid = policy_data.get("line_guid", "07564291-CBFE-4BBE-88D1-0548C88ACED4")
        else:
            # Fall back to original logic
            effective_date = datetime.now().date()
            expiration_date = datetime.now().date().replace(year=datetime.now().year + 1)
            
            if data.get("effective_date"):
                try:
                    effective_date = datetime.strptime(data["effective_date"], "%Y-%m-%d").date()
                except ValueError:
                    pass
                    
            if data.get("expiration_date"):
                try:
                    expiration_date = datetime.strptime(data["expiration_date"], "%Y-%m-%d").date()
                except ValueError:
                    pass
            
            state = data.get("state", "TX")
            line_of_business = data.get("line_of_business", "General Liability")
            line_guid = "07564291-CBFE-4BBE-88D1-0548C88ACED4"
        
        # Use hardcoded valid IMS GUIDs instead of all-zeros defaults
        result = {
            "submission_guid": None,  # Will be filled in by the caller
            "effective_date": effective_date,
            "expiration_date": expiration_date,
            "state": state,
            "line_guid": line_guid,
            "status_id": 1,  # Default status (New)
            "billing_type_id": 1,  # Default billing type (Agency Bill)
            "line_of_business": line_of_business,
            # Add hardcoded location GUIDs
            "quoting_location_guid": "C5C006BB-6437-42F3-95D4-C090ADD3B37D",
            "issuing_location_guid": "C5C006BB-6437-42F3-95D4-C090ADD3B37D",
            "company_location_guid": "DF35D4C7-C663-4974-A886-A1E18D3C9618"
        }
        
        # Extract line of business information if available
        if data.get("line_of_business"):
            # In a real implementation, you would look up the GUID based on line of business name
            # result["line_guid"] = self._get_line_guid(data["line_of_business"])
            pass
        
        return result
    
    def _extract_premium_data(self, transaction: Transaction) -> Dict[str, Any]:
        """Extract premium data from transaction"""
        # This method would extract and transform premium data from the transaction
        # to match IMS API requirements
        
        # Use processed_data if available (from transformers), otherwise use parsed_data
        data = transaction.processed_data or transaction.parsed_data
        if not data:
            raise ValueError("Cannot extract premium data: Missing data")
        
        # Check if we have transformed data
        if transaction.processed_data and "premium_data" in transaction.processed_data:
            # Use the transformed premium data
            premium_data = transaction.processed_data["premium_data"]
            total_premium = premium_data.get("gross_premium", 0.0)
            
            # Get coverages from transformed data
            coverages = []
            if "coverages" in transaction.processed_data:
                for coverage in transaction.processed_data["coverages"]:
                    coverages.append({
                        "type": coverage.get("coverage_type", "Unknown"),
                        "premium": coverage.get("premium", total_premium),
                        "limit": coverage.get("limit_occurrence"),
                        "limit_aggregate": coverage.get("limit_aggregate"),
                        "deductible": coverage.get("deductible")
                    })
        else:
            # Fall back to original logic
            total_premium = 0.0
            
            # Extract premium amount
            if data.get("premium"):
                try:
                    total_premium = float(data["premium"])
                except (ValueError, TypeError):
                    pass
            
            # Extract coverages if available
            coverages = []
            if "coverages" in data and isinstance(data["coverages"], list):
                for coverage in data["coverages"]:
                    if coverage.get("premium"):
                        try:
                            cov_premium = float(coverage["premium"])
                            coverages.append({
                                "type": coverage.get("type", "Unknown"),
                                "premium": cov_premium,
                                "limit": coverage.get("limit"),
                                "deductible": coverage.get("deductible")
                            })
                        except (ValueError, TypeError):
                            pass
        
        return {
            "total_premium": total_premium,
            "coverages": coverages
        }
    
    def _should_use_excel_rater(self, transaction: Transaction) -> bool:
        """Determine if the transaction should use Excel rater"""
        # Check if program is configured to use Excel rater
        # Default to manual rating (direct premium pass-through) for most programs
        
        if not transaction.parsed_data:
            return False
            
        # Check if transaction explicitly requests Excel rater
        use_excel_rater = transaction.parsed_data.get("use_excel_rater", False)
        
        # Only use Excel rater if explicitly requested AND template exists
        if use_excel_rater:
            return self._get_excel_template(transaction) is not None
            
        # Default: Use manual rating (direct premium pass-through)
        return False
    
    def _get_excel_template(self, transaction: Transaction) -> Optional[str]:
        """Get the Excel template for the transaction"""
        # This method would determine the appropriate Excel template
        # based on the line of business, program, etc.
        
        if not transaction.parsed_data:
            return None
        
        data = transaction.parsed_data
        line_of_business = data.get("line_of_business", "").lower().replace(" ", "_")
        program = data.get("program", "").lower().replace(" ", "_")
        state = data.get("state", "").upper()
        
        # Generate possible template paths
        template_paths = [
            f"{line_of_business}_{program}_{state}.xlsx",
            f"{line_of_business}_{program}.xlsx",
            f"{line_of_business}_{state}.xlsx",
            f"{line_of_business}.xlsx",
            "default.xlsx"
        ]
        
        # Check if any of the templates exist
        for template_name in template_paths:
            template_path = os.path.join(self.templates_dir, template_name)
            if os.path.exists(template_path):
                # Also set rater ID and factor set GUID based on template
                self._set_rater_info(transaction, template_name)
                return template_path
        
        # Create a sample template if none exists
        default_template = os.path.join(self.templates_dir, "default.xlsx")
        if not os.path.exists(default_template):
            self._create_sample_template(default_template)
            # Set default rater info
            self._set_rater_info(transaction, "default.xlsx")
            return default_template
            
        return None
    
    def _set_rater_info(self, transaction: Transaction, template_name: str) -> None:
        """Set rater ID and factor set GUID based on template"""
        # In a real implementation, you would look up these values based on the template
        transaction.ims_processing.excel_rater_id = 1
        transaction.ims_processing.factor_set_guid = "00000000-0000-0000-0000-000000000000"
    
    def _create_sample_template(self, template_path: str) -> None:
        """Create a sample Excel template"""
        # Create a basic Excel workbook with some sample data
        wb = Workbook()
        ws = wb.active
        ws.title = "Rate Calculator"
        
        # Add some headers
        headers = ["Policy_Number", "Effective_Date", "Expiration_Date", "Line_of_Business", 
                   "State", "Limit", "Deductible", "Premium"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add some formulas for premium calculation
        ws.cell(row=2, column=8, value="=IF(F2>0,F2*0.05,0)")  # Simple premium calculation
        
        # Save the workbook
        wb.save(template_path)
        logger.info(f"Created sample Excel template at {template_path}")
    
    def _populate_excel_template(self, template_path: str, transaction: Transaction) -> str:
        """Populate Excel template with transaction data"""
        if not transaction.parsed_data:
            raise ValueError("Cannot populate Excel template: Missing parsed data")
        
        data = transaction.parsed_data
        
        # Create a copy of the template in a temp file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_path = temp_file.name
        
        # Load the template
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Find headers
        headers = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers[header] = col
        
        # Populate basic data (row 2)
        if "Policy_Number" in headers and data.get("policy_number"):
            ws.cell(row=2, column=headers["Policy_Number"], value=data["policy_number"])
            
        if "Effective_Date" in headers and data.get("effective_date"):
            try:
                eff_date = datetime.strptime(data["effective_date"], "%Y-%m-%d").date()
                ws.cell(row=2, column=headers["Effective_Date"], value=eff_date)
            except ValueError:
                pass
                
        if "Expiration_Date" in headers and data.get("expiration_date"):
            try:
                exp_date = datetime.strptime(data["expiration_date"], "%Y-%m-%d").date()
                ws.cell(row=2, column=headers["Expiration_Date"], value=exp_date)
            except ValueError:
                pass
                
        if "Line_of_Business" in headers and data.get("line_of_business"):
            ws.cell(row=2, column=headers["Line_of_Business"], value=data["line_of_business"])
            
        if "State" in headers and data.get("state"):
            ws.cell(row=2, column=headers["State"], value=data["state"])
        
        # Populate coverage data
        if "coverages" in data and isinstance(data["coverages"], list) and data["coverages"]:
            coverage = data["coverages"][0]  # Just use the first coverage for now
            
            if "Limit" in headers and coverage.get("limit"):
                ws.cell(row=2, column=headers["Limit"], value=float(coverage["limit"]))
                
            if "Deductible" in headers and coverage.get("deductible"):
                ws.cell(row=2, column=headers["Deductible"], value=float(coverage["deductible"]))
        
        # Save the populated workbook
        wb.save(temp_path)
        
        return temp_path
    
