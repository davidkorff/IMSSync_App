"""
Simplified Triton transaction processor
Handles all Triton transaction types with clear, linear flow
"""

import logging
from typing import Dict, Any, Optional, Tuple
from datetime import datetime, date
import json

logger = logging.getLogger(__name__)


class TritonError(Exception):
    """Custom exception for Triton processing errors with context"""
    def __init__(self, stage: str, message: str, details: dict = None):
        self.stage = stage
        self.message = message
        self.details = details or {}
        super().__init__(f"[{stage}] {message}")


class TritonProcessor:
    """
    Processes Triton transactions with a simple, linear flow:
    1. Validate incoming data
    2. Transform to IMS format
    3. Call IMS APIs
    4. Return result
    """
    
    def __init__(self, ims_client, config: dict):
        self.ims = ims_client
        self.config = config
        
    def process_transaction(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Main entry point for processing any Triton transaction
        """
        transaction_type = data.get('transaction_type', '').lower()
        transaction_id = data.get('transaction_id', 'unknown')
        
        logger.info(f"TRITON_TRANSACTION_START", extra={
            'transaction_id': transaction_id,
            'transaction_type': transaction_type,
            'policy_number': data.get('policy_number')
        })
        
        try:
            # Route to appropriate handler
            # Normalize transaction types
            if transaction_type in ['binding', 'new', 'new business', 'new_business']:
                return self.process_binding(data)
            elif transaction_type == 'cancellation':
                return self.process_cancellation(data)
            elif transaction_type in ['endorsement', 'midterm_endorsement']:
                return self.process_endorsement(data)
            elif transaction_type == 'reinstatement':
                return self.process_reinstatement(data)
            else:
                raise TritonError(
                    'VALIDATION',
                    f'Unknown transaction type: {transaction_type}',
                    {'transaction_type': transaction_type}
                )
                
        except TritonError:
            raise
        except Exception as e:
            logger.exception(f"Unexpected error processing transaction {transaction_id}")
            raise TritonError(
                'PROCESSING',
                f'Unexpected error: {str(e)}',
                {'transaction_id': transaction_id, 'original_error': str(e)}
            )
    
    def process_binding(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process a policy binding transaction
        Simple flow: Insured → Submission → Quote → Premium/Rating → Bind
        """
        transaction_id = data.get('transaction_id', 'unknown')
        
        try:
            # 1. Transform and validate data
            logger.info("TRITON_BINDING_TRANSFORM", extra={'transaction_id': transaction_id})
            ims_data = self._transform_binding_data(data)
            
            # 2. Create or find insured
            logger.info("TRITON_BINDING_INSURED", extra={
                'transaction_id': transaction_id,
                'insured_name': ims_data['insured']['name']
            })
            insured_guid = self.ims.find_or_create_insured(ims_data['insured'])
            
            # 3. Create submission
            logger.info("TRITON_BINDING_SUBMISSION", extra={'transaction_id': transaction_id})
            submission_data = {
                'insured_guid': insured_guid,
                'producer_guid': ims_data['producer_guid'],
                'underwriter_guid': ims_data.get('underwriter_guid', self.config['defaults']['underwriter_guid']),
                'submission_date': ims_data['submission_date']
            }
            submission_guid = self.ims.create_submission(submission_data)
            
            # 4. Create quote with all required fields
            logger.info("TRITON_BINDING_QUOTE", extra={'transaction_id': transaction_id})
            quote_data = {
                'submission_guid': submission_guid,
                'effective_date': ims_data['effective_date'],
                'expiration_date': ims_data['expiration_date'],
                'state': ims_data['state'],
                'line_guid': ims_data['line_guid'],
                'producer_guid': ims_data['producer_guid'],
                'location_guids': ims_data['location_guids'],
                'underwriter_guid': ims_data.get('underwriter_guid', self.config['defaults']['underwriter_guid']),
                'insured_business_type_id': ims_data['insured']['business_type_id']  # Pass the business type ID
            }
            quote_guid = self.ims.create_quote(quote_data)
            
            # 5. Handle rating - either Excel or direct premium
            use_excel_rating = self.config.get('use_excel_rating', True)
            option_id = None
            
            if use_excel_rating and self._should_use_excel_rating(data):
                logger.info("TRITON_BINDING_EXCEL_RATING", extra={'transaction_id': transaction_id})
                
                # Generate Excel file with all Triton data
                excel_data = self._generate_excel_from_triton(data, ims_data)
                
                # Get rater configuration
                rater_config = self._get_rater_config(ims_data['state'], ims_data['line_type'])
                
                # Import Excel rater (this rates AND stores the data)
                rating_result = self.ims.import_excel_rater(
                    quote_guid=quote_guid,
                    file_bytes=excel_data,
                    file_name=f"triton_{transaction_id}.xlsx",
                    rater_id=rater_config['rater_id'],
                    factor_set_guid=rater_config.get('factor_set_guid'),
                    apply_fees=True
                )
                
                if rating_result['success'] and rating_result.get('premiums'):
                    option_result = rating_result['premiums'][0]
                    option_id = option_result['quote_option_guid']
                    logger.info("Excel rating successful", extra={
                        'premium': option_result['premium_total'],
                        'option_id': option_id
                    })
                else:
                    raise TritonError('RATING', 'Excel rating failed', rating_result)
                    
                # Also save the raw data as a rating sheet for reference
                self.ims.save_rating_sheet(
                    quote_guid=quote_guid,
                    rater_id=rater_config['rater_id'],
                    file_bytes=excel_data,
                    file_name=f"triton_raw_{transaction_id}.xlsx"
                )
                
            else:
                # Direct premium pass-through
                logger.info("TRITON_BINDING_PREMIUM", extra={
                    'transaction_id': transaction_id,
                    'premium': ims_data['premium']
                })
                option_id = self.ims.add_quote_option(quote_guid)
                self.ims.add_premium(quote_guid, option_id, ims_data['premium'])
                
                # Still save the Triton data as JSON in a custom table
                self._store_triton_data(quote_guid, data)
            
            # 6. Bind policy
            logger.info("TRITON_BINDING_BIND", extra={'transaction_id': transaction_id})
            policy_number = self.ims.bind_quote(option_id)
            
            # 7. Try to get invoice number (may not be immediately available)
            invoice_number = None
            try:
                invoice_data = self.ims.get_latest_invoice(policy_number)
                invoice_number = invoice_data.get('invoice_number') if invoice_data else None
            except:
                logger.info("Invoice not immediately available", extra={'policy_number': policy_number})
            
            # 8. Link external ID
            if data.get('transaction_id'):
                self.ims.update_external_id(quote_guid, data['transaction_id'], 'triton')
            
            logger.info("TRITON_BINDING_SUCCESS", extra={
                'transaction_id': transaction_id,
                'policy_number': policy_number,
                'quote_guid': quote_guid
            })
            
            return {
                'success': True,
                'transaction_id': transaction_id,
                'policy_number': policy_number,
                'quote_guid': str(quote_guid),
                'invoice_number': invoice_number,
                'ims_reference': str(quote_guid)
            }
            
        except Exception as e:
            logger.error(f"Error in binding process: {str(e)}", extra={
                'transaction_id': transaction_id,
                'error': str(e)
            })
            raise TritonError(
                'BINDING',
                str(e),
                {'transaction_id': transaction_id}
            )
    
    def process_cancellation(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process a policy cancellation
        """
        transaction_id = data.get('transaction_id', 'unknown')
        policy_number = data.get('policy_number', '')
        
        try:
            # 1. Get control number from policy number
            logger.info("TRITON_CANCELLATION_LOOKUP", extra={
                'transaction_id': transaction_id,
                'policy_number': policy_number
            })
            control_number = self.ims.get_control_number(policy_number)
            
            # 2. Process cancellation
            logger.info("TRITON_CANCELLATION_PROCESS", extra={
                'transaction_id': transaction_id,
                'control_number': control_number
            })
            
            cancellation_data = {
                'control_number': control_number,
                'cancellation_date': self._parse_date(data.get('cancellation_date')),
                'cancellation_reason': data.get('cancellation_reason', ''),
                'cancellation_reason_id': self._map_cancellation_reason(data.get('cancellation_reason')),
                'flat_cancel': data.get('flat_cancel', False),
                'user_guid': self.config['defaults']['user_guid']
            }
            
            result = self.ims.cancel_policy(cancellation_data)
            
            logger.info("TRITON_CANCELLATION_SUCCESS", extra={
                'transaction_id': transaction_id,
                'policy_number': policy_number,
                'control_number': control_number
            })
            
            return {
                'success': True,
                'transaction_id': transaction_id,
                'policy_number': policy_number,
                'message': result.get('message', 'Policy cancelled successfully'),
                'ims_reference': str(control_number)
            }
            
        except Exception as e:
            logger.error(f"Error in cancellation process: {str(e)}", extra={
                'transaction_id': transaction_id,
                'policy_number': policy_number,
                'error': str(e)
            })
            raise TritonError(
                'CANCELLATION',
                str(e),
                {'transaction_id': transaction_id, 'policy_number': policy_number}
            )
    
    def process_endorsement(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process a policy endorsement
        """
        transaction_id = data.get('transaction_id', 'unknown')
        policy_number = data.get('policy_number', '')
        
        try:
            # 1. Get control number
            logger.info("TRITON_ENDORSEMENT_LOOKUP", extra={
                'transaction_id': transaction_id,
                'policy_number': policy_number
            })
            control_number = self.ims.get_control_number(policy_number)
            
            # 2. Create endorsement
            logger.info("TRITON_ENDORSEMENT_CREATE", extra={
                'transaction_id': transaction_id,
                'control_number': control_number
            })
            
            endorsement = data.get('endorsement', {})
            endorsement_data = {
                'control_number': control_number,
                'effective_date': self._parse_date(endorsement.get('effective_from') or data.get('effective_date')),
                'description': endorsement.get('description', 'Policy endorsement via Triton'),
                'reason_id': self._map_endorsement_reason(endorsement.get('endorsement_code')),
                'user_guid': self.config['defaults']['user_guid']
            }
            
            result = self.ims.create_endorsement(endorsement_data)
            endorsement_guid = result.get('endorsement_quote_guid')
            
            # 3. Apply premium change if any
            premium_change = endorsement.get('premium', 0)
            if premium_change and endorsement_guid:
                logger.info("TRITON_ENDORSEMENT_PREMIUM", extra={
                    'transaction_id': transaction_id,
                    'premium_change': premium_change
                })
                self.ims.add_endorsement_premium(endorsement_guid, premium_change)
            
            logger.info("TRITON_ENDORSEMENT_SUCCESS", extra={
                'transaction_id': transaction_id,
                'policy_number': policy_number,
                'endorsement_guid': endorsement_guid
            })
            
            return {
                'success': True,
                'transaction_id': transaction_id,
                'policy_number': policy_number,
                'endorsement_guid': str(endorsement_guid) if endorsement_guid else None,
                'message': 'Endorsement created successfully',
                'ims_reference': str(endorsement_guid) if endorsement_guid else str(control_number)
            }
            
        except Exception as e:
            logger.error(f"Error in endorsement process: {str(e)}", extra={
                'transaction_id': transaction_id,
                'policy_number': policy_number,
                'error': str(e)
            })
            raise TritonError(
                'ENDORSEMENT',
                str(e),
                {'transaction_id': transaction_id, 'policy_number': policy_number}
            )
    
    def process_reinstatement(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process a policy reinstatement
        """
        transaction_id = data.get('transaction_id', 'unknown')
        policy_number = data.get('policy_number', '')
        
        try:
            # 1. Get control number
            logger.info("TRITON_REINSTATEMENT_LOOKUP", extra={
                'transaction_id': transaction_id,
                'policy_number': policy_number
            })
            control_number = self.ims.get_control_number(policy_number)
            
            # 2. Process reinstatement
            logger.info("TRITON_REINSTATEMENT_PROCESS", extra={
                'transaction_id': transaction_id,
                'control_number': control_number
            })
            
            reinstatement_data = {
                'control_number': control_number,
                'reinstatement_date': self._parse_date(data.get('reinstatement_date') or data.get('effective_date')),
                'reason_id': self._map_reinstatement_reason(data.get('reason')),
                'comments': data.get('comments', 'Policy reinstated via Triton'),
                'payment_received': data.get('payment_amount'),
                'check_number': data.get('check_number'),
                'user_guid': self.config['defaults']['user_guid']
            }
            
            result = self.ims.reinstate_policy(reinstatement_data)
            
            logger.info("TRITON_REINSTATEMENT_SUCCESS", extra={
                'transaction_id': transaction_id,
                'policy_number': policy_number,
                'invoice_number': result.get('invoice_number')
            })
            
            return {
                'success': True,
                'transaction_id': transaction_id,
                'policy_number': policy_number,
                'invoice_number': result.get('invoice_number'),
                'message': 'Policy reinstated successfully',
                'ims_reference': str(control_number)
            }
            
        except Exception as e:
            logger.error(f"Error in reinstatement process: {str(e)}", extra={
                'transaction_id': transaction_id,
                'policy_number': policy_number,
                'error': str(e)
            })
            raise TritonError(
                'REINSTATEMENT',
                str(e),
                {'transaction_id': transaction_id, 'policy_number': policy_number}
            )
    
    def _transform_binding_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Transform Triton binding data to IMS format
        Handles both nested (account/producer) and flat (insured_name/producer_name) structures
        """
        # Check if this is flat structure (like TEST.json)
        if 'insured_name' in data and 'account' not in data:
            return self._transform_flat_binding_data(data)
        
        # Extract account/insured data (nested structure)
        account = data.get('account', {})
        
        # Map business type from account data
        business_type = account.get('business_type', '').lower()
        business_type_mapping = {
            'individual': 4,
            'partnership': 2,
            'limited partnership': 3,
            'llc': 9,
            'llp': 9,
            'llc/llp': 9,
            'joint venture': 10,
            'trust': 11,
            'corporation': 13,
            'other': 5
        }
        business_type_id = business_type_mapping.get(business_type, 5)  # Default to 5 (Other)
        logger.info(f"Mapped business type '{business_type}' to ID {business_type_id}")
        
        insured_data = {
            'name': account.get('name', ''),
            'tax_id': account.get('id'),
            'business_type_id': business_type_id,
            'address': account.get('street_1', ''),
            'city': account.get('city', ''),
            'state': account.get('state', ''),
            'zip': account.get('zip', '')
        }
        
        # Determine line of business
        line_guid = self._determine_line_guid(data)
        
        # Get producer GUID
        producer_guid = self._get_producer_guid(data)
        
        # Extract dates
        effective_date = self._parse_date(data.get('effective_date'))
        expiration_date = self._parse_date(data.get('expiration_date'))
        
        # Extract premium
        premium_data = data.get('premium', {})
        total_premium = premium_data.get('annual_premium', 0) if isinstance(premium_data, dict) else float(premium_data or 0)
        
        # Build location GUIDs
        location_guids = {
            'issuing': self.config['locations']['issuing_guid'],
            'company': self.config['locations']['company_guid'],
            'quoting': self.config['locations']['quoting_guid']
        }
        
        # Determine line type for rating
        line_type = 'primary'
        if line_guid == self.config['lines'].get('excess'):
            line_type = 'excess'
        
        return {
            'insured': insured_data,
            'producer_guid': producer_guid,
            'line_guid': line_guid,
            'line_type': line_type,
            'effective_date': effective_date,
            'expiration_date': expiration_date,
            'submission_date': datetime.now().date(),
            'state': account.get('state', 'TX'),
            'premium': total_premium,
            'location_guids': location_guids
        }
    
    def _determine_line_guid(self, data: Dict[str, Any]) -> str:
        """
        Determine the appropriate line GUID based on coverage type
        """
        # Check program info
        program = data.get('program', {})
        if isinstance(program, dict):
            program_name = program.get('name', '').lower()
            if 'excess' in program_name:
                return self.config['lines']['excess']
        
        # Check exposures/coverages
        exposures = data.get('exposures', [])
        for exposure in exposures:
            if isinstance(exposure, dict):
                coverage = exposure.get('coverage_name', '').lower()
                if 'excess' in coverage:
                    return self.config['lines']['excess']
        
        # Default to primary
        return self.config['lines']['primary']
    
    def _get_producer_guid(self, data: Dict[str, Any]) -> str:
        """
        Get producer GUID from data or config
        """
        producer = data.get('producer', {})
        if isinstance(producer, dict):
            producer_name = producer.get('name', '')
            # Check mapping
            if producer_name in self.config['producers']['mapping']:
                return self.config['producers']['mapping'][producer_name]
        
        # Return default
        return self.config['producers']['default']
    
    def _parse_date(self, date_value: Any) -> date:
        """
        Parse date from various formats
        """
        if isinstance(date_value, date):
            return date_value
        
        if isinstance(date_value, datetime):
            return date_value.date()
        
        if isinstance(date_value, str):
            # Try common date formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d']:
                try:
                    return datetime.strptime(date_value, fmt).date()
                except ValueError:
                    continue
        
        # Default to today
        return datetime.now().date()
    
    def _map_cancellation_reason(self, reason: str) -> int:
        """Map Triton cancellation reason to IMS ID"""
        mapping = {
            'non-payment': 1,
            'request': 2,
            'underwriting': 3,
            'fraud': 4
        }
        return mapping.get(str(reason).lower(), 99)
    
    def _map_endorsement_reason(self, code: str) -> int:
        """Map Triton endorsement code to IMS ID"""
        mapping = {
            'add_coverage': 1,
            'remove_coverage': 2,
            'change_limit': 3,
            'add_location': 4
        }
        return mapping.get(str(code).lower(), 99)
    
    def _map_reinstatement_reason(self, reason: str) -> int:
        """Map reinstatement reason to IMS ID"""
        mapping = {
            'payment_received': 1,
            'error': 2,
            'appeal': 3
        }
        return mapping.get(str(reason).lower(), 99)
    
    def _should_use_excel_rating(self, data: Dict[str, Any]) -> bool:
        """
        Determine if Excel rating should be used based on data or config
        """
        # Check if explicitly requested in data
        if 'use_excel_rating' in data:
            return bool(data['use_excel_rating'])
        
        # Check if there's complex rating data that needs Excel
        has_complex_data = (
            'exposures' in data and len(data.get('exposures', [])) > 1 or
            'rating_factors' in data or
            'custom_fields' in data
        )
        
        return has_complex_data
    
    def _generate_excel_from_triton(self, triton_data: Dict[str, Any], ims_data: Dict[str, Any]) -> bytes:
        """
        Generate Excel file from Triton data for IMS rating
        This preserves ALL Triton data including fields IMS doesn't have
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            import io
            
            wb = Workbook()
            
            # Main rating sheet
            ws = wb.active
            ws.title = "Rating Data"
            
            # Headers for standard fields
            headers = [
                'Policy_Number', 'Effective_Date', 'Expiration_Date',
                'Insured_Name', 'State', 'Premium', 'Producer',
                'Program', 'Coverage_Type', 'Limit', 'Deductible'
            ]
            
            # Add headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Populate standard data
            row = 2
            account = triton_data.get('account', {})
            program = triton_data.get('program', {})
            
            ws.cell(row=row, column=1, value=triton_data.get('policy_number', ''))
            ws.cell(row=row, column=2, value=str(ims_data['effective_date']))
            ws.cell(row=row, column=3, value=str(ims_data['expiration_date']))
            ws.cell(row=row, column=4, value=account.get('name', ''))
            ws.cell(row=row, column=5, value=ims_data['state'])
            ws.cell(row=row, column=6, value=ims_data['premium'])
            ws.cell(row=row, column=7, value=triton_data.get('producer', {}).get('name', ''))
            ws.cell(row=row, column=8, value=program.get('name', ''))
            
            # Add exposures/coverages
            exposures = triton_data.get('exposures', [])
            if exposures:
                exposure = exposures[0]  # Primary exposure
                ws.cell(row=row, column=9, value=exposure.get('coverage_name', ''))
                
                # Handle limit
                limit = exposure.get('limit', {})
                if isinstance(limit, dict):
                    ws.cell(row=row, column=10, value=limit.get('value', 0))
                else:
                    ws.cell(row=row, column=10, value=limit)
                
                # Handle deductible
                deductible = exposure.get('deductible', 0)
                if isinstance(deductible, dict):
                    ws.cell(row=row, column=11, value=deductible.get('value', 0))
                else:
                    ws.cell(row=row, column=11, value=deductible)
            
            # Create a separate sheet for ALL Triton data (preserved as-is)
            ws_raw = wb.create_sheet("Triton Raw Data")
            ws_raw.cell(row=1, column=1, value="Field")
            ws_raw.cell(row=1, column=2, value="Value")
            
            # Flatten and store all Triton data
            row = 2
            for key, value in self._flatten_dict(triton_data).items():
                ws_raw.cell(row=row, column=1, value=key)
                ws_raw.cell(row=row, column=2, value=str(value))
                row += 1
            
            # Convert to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
            
        except ImportError:
            logger.error("openpyxl not installed - cannot generate Excel")
            # Fallback: return empty Excel-like data
            return b''
    
    def _flatten_dict(self, d: Dict[str, Any], parent_key: str = '', sep: str = '.') -> Dict[str, Any]:
        """
        Flatten nested dictionary for Excel storage
        """
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                for i, item in enumerate(v):
                    if isinstance(item, dict):
                        items.extend(self._flatten_dict(item, f"{new_key}[{i}]", sep=sep).items())
                    else:
                        items.append((f"{new_key}[{i}]", item))
            else:
                items.append((new_key, v))
        return dict(items)
    
    def _get_rater_config(self, state: str, line_type: str) -> Dict[str, Any]:
        """
        Get rater configuration based on state and line type
        """
        # Check for state-specific rater
        state_key = f"{line_type}_{state}"
        if state_key in self.config.get('raters', {}):
            return self.config['raters'][state_key]
        
        # Check for line-specific rater
        if line_type in self.config.get('raters', {}):
            return self.config['raters'][line_type]
        
        # Return default rater
        return self.config.get('raters', {}).get('default', {
            'rater_id': 1,
            'factor_set_guid': None
        })
    
    def _transform_flat_binding_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Transform flat Triton data (like TEST.json) to IMS format
        """
        # Map business type string to IMS ID
        business_type = data.get('business_type', '').lower()
        business_type_mapping = {
            'individual': 4,
            'partnership': 2,
            'limited partnership': 3,
            'llc': 9,
            'llp': 9,
            'llc/llp': 9,
            'joint venture': 10,
            'trust': 11,
            'corporation': 13,
            'other': 5
        }
        business_type_id = business_type_mapping.get(business_type, 5)  # Default to 5 (Other)
        logger.info(f"Mapped business type '{business_type}' to ID {business_type_id}")
        
        # Build insured data from flat structure
        insured_data = {
            'name': data.get('insured_name', ''),
            'tax_id': data.get('tax_id') or data.get('fein') or data.get('insured_fein'),
            'business_type_id': business_type_id,
            'address': data.get('address_1', ''),
            'city': data.get('city', '') or data.get('insured_city', ''),
            'state': data.get('state', '') or data.get('insured_state', ''),
            'zip': data.get('zip', '') or data.get('insured_zip', '')
        }
        
        # Determine line of business from program name
        program_name = data.get('program_name', '').lower()
        if 'excess' in program_name:
            line_guid = self.config['lines']['excess']
            line_type = 'excess'
        else:
            line_guid = self.config['lines']['primary']
            line_type = 'primary'
        
        # Get producer GUID
        producer_name = data.get('producer_name', '')
        producer_guid = self.config['producers']['mapping'].get(producer_name, self.config['producers']['default'])
        
        # Parse dates
        effective_date = self._parse_date(data.get('effective_date'))
        expiration_date = self._parse_date(data.get('expiration_date'))
        
        # Get premium
        total_premium = float(data.get('gross_premium', 0) or data.get('net_premium', 0))
        
        # Build location GUIDs
        location_guids = {
            'issuing': self.config['locations']['issuing'],
            'company': self.config['locations']['company'],
            'quoting': self.config['locations']['quoting']
        }
        
        return {
            'insured': insured_data,
            'producer_guid': producer_guid,
            'line_guid': line_guid,
            'line_type': line_type,
            'effective_date': effective_date,
            'expiration_date': expiration_date,
            'submission_date': self._parse_date(data.get('bound_date')) or datetime.now().date(),
            'state': data.get('state') or data.get('insured_state', 'TX'),
            'premium': total_premium,
            'location_guids': location_guids
        }
    
    def _store_triton_data(self, quote_guid: str, data: Dict[str, Any]) -> None:
        """
        Store Triton data in IMS using custom stored procedure
        This is used when not using Excel rating
        """
        try:
            # Convert data to JSON string
            json_data = json.dumps(data)
            
            # Call custom stored procedure to store data
            self.ims.execute_command(
                'StoreTritonData',
                [
                    'QuoteGUID', str(quote_guid),
                    'SourceSystem', 'triton',
                    'TransactionID', data.get('transaction_id', ''),
                    'JsonData', json_data
                ]
            )
            
            logger.info("Stored Triton data in IMS", extra={
                'quote_guid': quote_guid,
                'transaction_id': data.get('transaction_id')
            })
            
        except Exception as e:
            # Log but don't fail the transaction
            logger.warning(f"Could not store Triton data: {str(e)}", extra={
                'quote_guid': quote_guid,
                'error': str(e)
            })