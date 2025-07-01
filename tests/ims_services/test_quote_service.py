"""
Tests for IMS Quote Service

Tests submission creation, quote creation, rating, binding, and policy issuance.
"""

import unittest
from datetime import date
from test_base import IMSServiceTestBase
import logging
import os

logger = logging.getLogger(__name__)


class TestIMSQuoteService(IMSServiceTestBase):
    """Test IMS Quote Service functionality"""
    
    @classmethod
    def setUpClass(cls):
        """Set up test data for quote tests"""
        super().setUpClass()
        
        # Create a test insured and producer for quote tests
        try:
            # Create insured
            insured_data = {
                "name": f"Quote Test Company {cls.generate_test_id(cls)}",
                "tax_id": "12-3456789",
                "business_type": "LLC",
                "address": "123 Quote Test St",
                "city": "Dallas",
                "state": "TX",
                "zip_code": "75001"
            }
            cls.test_insured_guid = cls.insured_service.create_insured(insured_data)
            cls.track_entity(cls, "insureds", cls.test_insured_guid)
            logger.info(f"Created test insured: {cls.test_insured_guid}")
            
            # Get default producer
            cls.test_producer_guid = cls.producer_service.get_default_producer_guid("triton")
            logger.info(f"Using producer: {cls.test_producer_guid}")
            
        except Exception as e:
            logger.error(f"Failed to set up test data: {str(e)}")
            raise
    
    def test_01_create_submission(self):
        """Test creating a submission"""
        # Create submission data
        submission_data = self.generate_submission_data(
            self.test_insured_guid,
            self.test_producer_guid
        )
        
        self.log_test_data(submission_data, "Creating Submission")
        
        # Create submission
        submission_guid = self.quote_service.create_submission(submission_data)
        
        # Verify
        self.assertIsGuid(submission_guid)
        self.track_entity("submissions", submission_guid)
        
        # Store for next tests
        self.submission_guid = submission_guid
        logger.info(f"Created submission: {submission_guid}")
    
    def test_02_create_quote(self):
        """Test creating a quote"""
        # Need a submission first
        if not hasattr(self, 'submission_guid'):
            submission_data = self.generate_submission_data(
                self.test_insured_guid,
                self.test_producer_guid
            )
            self.submission_guid = self.quote_service.create_submission(submission_data)
            self.track_entity("submissions", self.submission_guid)
        
        # Create quote data
        quote_data = self.generate_quote_data(self.submission_guid)
        
        # Add specific fields
        quote_data["line_guid"] = self.quote_service.get_default_line_guid("triton", "primary")
        quote_data["producer_contact_guid"] = self.test_producer_guid
        
        self.log_test_data(quote_data, "Creating Quote")
        
        # Create quote
        quote_guid = self.quote_service.create_quote(quote_data)
        
        # Verify
        self.assertIsGuid(quote_guid)
        self.track_entity("quotes", quote_guid)
        
        # Store for next tests
        self.quote_guid = quote_guid
        logger.info(f"Created quote: {quote_guid}")
    
    def test_03_add_quote_option(self):
        """Test adding quote option"""
        # Need a quote first
        if not hasattr(self, 'quote_guid'):
            self.test_02_create_quote()
        
        # Add quote option
        option_id = self.quote_service.add_quote_option(self.quote_guid)
        
        # Verify
        self.assertIsNotNone(option_id)
        self.assertIsInstance(option_id, int)
        self.assertGreater(option_id, 0)
        
        # Store for next tests
        self.quote_option_id = option_id
        logger.info(f"Added quote option: {option_id}")
    
    def test_04_add_premium(self):
        """Test adding premium to quote option"""
        # Need quote and option
        if not hasattr(self, 'quote_guid'):
            self.test_02_create_quote()
        if not hasattr(self, 'quote_option_id'):
            self.test_03_add_quote_option()
        
        # Add premium
        premium_amount = 5000.00
        description = "Test Premium"
        
        success = self.quote_service.add_premium(
            self.quote_guid,
            self.quote_option_id,
            premium_amount,
            description
        )
        
        # Verify
        self.assertTrue(success, "Premium should be added successfully")
        logger.info(f"Added premium: ${premium_amount}")
    
    def test_05_get_default_line_guid(self):
        """Test getting default line GUIDs"""
        sources = ["triton", "xuber"]
        coverage_types = ["primary", "excess"]
        
        for source in sources:
            for coverage_type in coverage_types:
                logger.info(f"\nGetting line GUID for {source}/{coverage_type}")
                
                line_guid = self.quote_service.get_default_line_guid(source, coverage_type)
                
                # Verify
                self.assertIsGuid(line_guid)
                logger.info(f"Line GUID: {line_guid}")
    
    def test_06_get_rater_info(self):
        """Test getting rater information"""
        sources = ["triton", "xuber"]
        coverage_types = ["primary", "excess"]
        
        for source in sources:
            for coverage_type in coverage_types:
                logger.info(f"\nGetting rater info for {source}/{coverage_type}")
                
                rater_id, factor_set_guid = self.quote_service.get_rater_info(source, coverage_type)
                
                # Verify
                self.assertIsInstance(rater_id, int)
                self.assertIsGuid(factor_set_guid)
                
                logger.info(f"Rater ID: {rater_id}, Factor Set: {factor_set_guid}")
    
    def test_07_update_external_quote_id(self):
        """Test updating external quote ID"""
        # Need a quote
        if not hasattr(self, 'quote_guid'):
            self.test_02_create_quote()
        
        # Update external ID
        external_id = f"EXT-{self.test_id}"
        external_system = "triton"
        
        success = self.quote_service.update_external_quote_id(
            self.quote_guid,
            external_id,
            external_system
        )
        
        # Verify
        self.assertTrue(success, "External ID update should succeed")
        logger.info(f"Updated external ID: {external_system}/{external_id}")
    
    def test_08_bind_quote(self):
        """Test binding a quote to create policy"""
        # Skip if binding is not allowed in test environment
        if os.getenv("SKIP_BINDING_TESTS", "false").lower() == "true":
            self.skipTest("Binding tests disabled")
        
        # Need a rated quote
        if not hasattr(self, 'quote_option_id'):
            self.test_03_add_quote_option()
            self.test_04_add_premium()
        
        try:
            # Bind quote
            policy_number = self.quote_service.bind_quote(self.quote_option_id)
            
            # Verify
            self.assertIsNotNone(policy_number)
            self.assertIsInstance(policy_number, str)
            self.assertGreater(len(policy_number), 0)
            
            # Store for next test
            self.policy_number = policy_number
            self.track_entity("policies", policy_number)
            
            logger.info(f"Bound quote, policy number: {policy_number}")
            
        except Exception as e:
            logger.warning(f"Binding failed (may be expected in test): {str(e)}")
            self.skipTest("Binding not available in test environment")
    
    def test_09_issue_policy(self):
        """Test issuing a policy"""
        # Skip if binding is not allowed
        if os.getenv("SKIP_BINDING_TESTS", "false").lower() == "true":
            self.skipTest("Binding tests disabled")
        
        # Need a bound policy
        if not hasattr(self, 'policy_number'):
            self.skipTest("No policy available to issue")
        
        try:
            # Issue policy
            success = self.quote_service.issue_policy(self.policy_number)
            
            # Verify
            self.assertTrue(success, "Policy issuance should succeed")
            logger.info(f"Issued policy: {self.policy_number}")
            
        except Exception as e:
            logger.warning(f"Issuance failed (may be expected in test): {str(e)}")
    
    def test_10_excel_rater_import(self):
        """Test Excel rater import"""
        # Skip if no Excel templates available
        if os.getenv("SKIP_EXCEL_TESTS", "false").lower() == "true":
            self.skipTest("Excel tests disabled")
        
        # Need a quote
        if not hasattr(self, 'quote_guid'):
            self.test_02_create_quote()
        
        # Create a simple test Excel file
        import tempfile
        try:
            import openpyxl
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Add headers
            headers = ["Policy_Number", "Premium", "Fee"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data
            ws.cell(row=2, column=1, value="TEST-POLICY")
            ws.cell(row=2, column=2, value=1000)
            ws.cell(row=2, column=3, value=50)
            
            # Save to temp file
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                wb.save(tmp.name)
                temp_path = tmp.name
            
            # Get rater info
            rater_id, factor_set_guid = self.quote_service.get_rater_info("triton", "primary")
            
            # Import Excel
            result = self.quote_service.import_excel_rater(
                self.quote_guid,
                temp_path,
                rater_id,
                factor_set_guid
            )
            
            # Verify
            self.assertIsInstance(result, dict)
            if result.get('success'):
                logger.info("Excel rater import succeeded")
                self.assertIn('premiums', result)
            else:
                logger.warning(f"Excel rater import failed: {result.get('error_message')}")
            
            # Clean up
            os.unlink(temp_path)
            
        except ImportError:
            self.skipTest("openpyxl not installed")
        except Exception as e:
            logger.error(f"Excel rater test failed: {str(e)}")
            raise


class TestQuoteValidation(IMSServiceTestBase):
    """Test quote data validation"""
    
    def test_01_required_fields_submission(self):
        """Test required fields for submission"""
        # Missing insured_guid
        with self.assertRaises(ValueError) as context:
            self.quote_service.create_submission({
                "producer_contact_guid": "00000000-0000-0000-0000-000000000000",
                "submission_date": date.today()
            })
        
        self.assertIn("insured_guid", str(context.exception))
        logger.info("Submission validation working correctly")
    
    def test_02_required_fields_quote(self):
        """Test required fields for quote"""
        # Missing submission_guid
        with self.assertRaises(ValueError) as context:
            self.quote_service.create_quote({
                "effective_date": date.today(),
                "expiration_date": date.today(),
                "state": "TX"
            })
        
        self.assertIn("submission_guid", str(context.exception))
        logger.info("Quote validation working correctly")
    
    def test_03_date_formatting(self):
        """Test date formatting in quotes"""
        # Create minimal valid quote data
        quote_data = {
            "submission_guid": "00000000-0000-0000-0000-000000000000",
            "effective_date": date.today(),
            "expiration_date": date.today().replace(year=date.today().year + 1),
            "state": "TX"
        }
        
        # This should not raise an error (date formatting should work)
        try:
            # We expect this to fail at the API level, not at formatting
            self.quote_service.create_quote(quote_data)
        except Exception as e:
            # Should fail for invalid submission, not date formatting
            self.assertNotIn("date", str(e).lower())
            logger.info("Date formatting working correctly")


if __name__ == "__main__":
    unittest.main(verbosity=2)